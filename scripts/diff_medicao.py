"""
diff_medicao.py — comparar duas medições xlsx por valor de célula.

Layer 2 de validação: cmp (Layer 1) detecta artefatos de ambiente (timestamps
zip, dcterms:modified, sharedStrings reorder); este script compara o conteúdo
real, célula por célula, da planilha primária do workbook.

Uso:
    python scripts/diff_medicao.py <baseline.xlsx> <candidato.xlsx>
Saída:
    "OK" + exit 0 quando zero diferenças.
    Lista de diffs (até 50) + exit 1 quando há diferenças de dados.
"""

import sys
from pathlib import Path

from openpyxl import load_workbook

CAP = 50


def comparar_celulas(caminho_a: str, caminho_b: str) -> list[tuple]:
    wb_a = load_workbook(caminho_a, read_only=True, data_only=True)
    wb_b = load_workbook(caminho_b, read_only=True, data_only=True)
    try:
        ws_a = wb_a[wb_a.sheetnames[0]]
        ws_b = wb_b[wb_b.sheetnames[0]]
        diffs = []
        rows_a = ws_a.iter_rows(values_only=True)
        rows_b = ws_b.iter_rows(values_only=True)
        for linha, (row_a, row_b) in enumerate(zip(rows_a, rows_b, strict=False), start=1):
            if row_a == row_b:
                continue
            largura = max(len(row_a), len(row_b))
            for col in range(largura):
                va = row_a[col] if col < len(row_a) else None
                vb = row_b[col] if col < len(row_b) else None
                if va != vb:
                    diffs.append((linha, col + 1, va, vb))
                    if len(diffs) >= CAP:
                        return diffs
        return diffs
    finally:
        wb_a.close()
        wb_b.close()


def main(argv: list[str]) -> int:
    if len(argv) != 3:
        print("uso: diff_medicao.py <baseline.xlsx> <candidato.xlsx>", file=sys.stderr)
        return 2
    a, b = Path(argv[1]), Path(argv[2])
    if not a.exists() or not b.exists():
        print(f"arquivo inexistente: {a if not a.exists() else b}", file=sys.stderr)
        return 2
    diffs = comparar_celulas(str(a), str(b))
    if not diffs:
        print("OK")
        return 0
    print(f"DIFF: {len(diffs)} célula(s) divergente(s) (cap={CAP})")
    for linha, col, va, vb in diffs:
        print(f"  L{linha} C{col}: baseline={va!r} candidato={vb!r}")
    return 1


if __name__ == "__main__":
    sys.exit(main(sys.argv))
