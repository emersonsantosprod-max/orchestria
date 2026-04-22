"""Comparação célula-a-célula entre medição original e processada.

Constantes de colunas, normalização por tipo e classificação de erros.
"""

import re
from collections import Counter
from datetime import time as dt_time

import openpyxl

NOME_ABA = 'Frequencia'

COL_DATA           = 0
COL_RE             = 1
COL_SUPERVISOR     = 2
COL_ENCARREGADO    = 3
COL_RONDA          = 4
COL_NOME           = 5
COL_SG_FUNCAO      = 6
COL_UNIDADE        = 7
COL_MD_COBRANCA    = 8
COL_SITUACAO       = 9
COL_UNID_ORIGEM    = 10
COL_TAG            = 11
COL_ID_JORNADA     = 12
COL_ENTRADA        = 13
COL_SAIDA          = 14
COL_APOIO_CALC     = 15
COL_DESC_DESCANSO  = 16
COL_DIF_TURNO      = 17
COL_DESCONTOS      = 18
COL_HR_TRAB        = 19
COL_PCT_COBR       = 20
COL_TIPO_RATEIO    = 21
COL_OBSERVACAO     = 22
COL_TOTAL_DESC     = 23
COL_HH_MEDIDO      = 24
COL_HISTOGRAMA     = 25

NOMES_COLUNAS = {
    0:  'Data',
    1:  'RE',
    2:  'Supervisor',
    3:  'Encarregado',
    4:  'Ronda',
    5:  'Nome',
    6:  'Sg Função',
    7:  'Unidade',
    8:  'MD Cobranca',
    9:  'Situação',
    10: 'Unid. Origem',
    11: 'TAG',
    12: 'ID Jornada',
    13: 'Entrada',
    14: 'Saída',
    15: 'Apoio Calculo Descanso',
    16: 'Desconto Descanso',
    17: 'Dif Turno',
    18: 'Descontos',
    19: 'Hr Trabalhadas',
    20: '% Cobrança',
    21: 'Tipo Rateio',
    22: 'Observação',
    23: 'Total Descontos',
    24: 'HH Medido',
    25: 'Histograma',
}

COLUNAS_MODIFICADAS = {COL_DESCONTOS, COL_OBSERVACAO}
COLUNAS_FORMULA_DEPENDENTE = {COL_HR_TRAB, COL_TOTAL_DESC, COL_HH_MEDIDO, COL_HISTOGRAMA}
COLUNAS_TEMPO = {COL_ENTRADA, COL_SAIDA, COL_APOIO_CALC, COL_DESC_DESCANSO, COL_DIF_TURNO, COL_DESCONTOS}

ERRO_CRITICO     = 'ERRO_CRITICO'
ERRO_NUMERICO    = 'ERRO_NUMERICO'
ERRO_DATA        = 'ERRO_DATA'
ERRO_MAPEAMENTO  = 'ERRO_MAPEAMENTO'
ERRO_ESTRUTURAL  = 'ERRO_ESTRUTURAL'
ERRO_FORMATO     = 'ERRO_FORMATO'
CRITERIO_ND      = 'CRITÉRIO NÃO DEFINIDO'
INFO_MODIFICACAO = 'INFO_MODIFICACAO'

_RE_HHMMSS = re.compile(r'^(\d{1,2}):(\d{2})(?::(\d{2}))?$')


def normalizar_time(valor) -> str:
    if valor is None or valor == '':
        return ''
    if isinstance(valor, dt_time):
        return f"{valor.hour:02d}:{valor.minute:02d}"
    if isinstance(valor, str):
        s = valor.strip()
        m = _RE_HHMMSS.match(s)
        if m:
            return f"{int(m.group(1)):02d}:{int(m.group(2)):02d}"
        return s
    if hasattr(valor, 'hour'):
        return f"{valor.hour:02d}:{valor.minute:02d}"
    return str(valor).strip()


def normalizar_datetime(valor) -> str:
    if valor is None:
        return ''
    if hasattr(valor, 'strftime'):
        return valor.strftime('%d/%m/%Y')
    return str(valor).strip()


def normalizar_float(valor, decimais: int = 4) -> str:
    if valor is None:
        return ''
    try:
        return f"{float(valor):.{decimais}f}"
    except (TypeError, ValueError):
        return str(valor).strip()


def normalizar_valor(valor, col_idx: int) -> str:
    if valor is None:
        return ''

    if col_idx == COL_DATA:
        return normalizar_datetime(valor)

    if col_idx in COLUNAS_TEMPO:
        return normalizar_time(valor)

    if col_idx in {COL_HR_TRAB, COL_PCT_COBR, COL_TOTAL_DESC, COL_HH_MEDIDO}:
        return normalizar_float(valor, decimais=4)

    if col_idx == COL_HISTOGRAMA:
        return normalizar_float(valor, decimais=6)

    if col_idx == COL_ID_JORNADA:
        try:
            return str(int(float(valor)))
        except (TypeError, ValueError):
            pass

    return str(valor).strip()


def classificar_erro(col_idx: int, col18_mudou: bool) -> str:
    if col_idx in COLUNAS_MODIFICADAS:
        return INFO_MODIFICACAO

    if col_idx in COLUNAS_FORMULA_DEPENDENTE:
        return ERRO_NUMERICO if col18_mudou else ERRO_CRITICO

    if col_idx == COL_DATA:
        return ERRO_DATA

    if col_idx == COL_RE:
        return ERRO_MAPEAMENTO

    return ERRO_CRITICO


def contexto_erro(tipo: str, col_idx: int, col18_mudou: bool) -> str:
    if tipo == INFO_MODIFICACAO:
        return 'Coluna modificada pelo pipeline — diferença esperada'
    if tipo == ERRO_NUMERICO and col_idx in COLUNAS_FORMULA_DEPENDENTE:
        return (
            'Fórmula Excel recalculada após atualização de col 18 (Descontos) — '
            'derivação esperada'
        )
    if tipo == ERRO_DATA:
        return 'Divergência em campo de data — não deve ser modificado pelo pipeline'
    if tipo == ERRO_MAPEAMENTO:
        return 'Divergência em campo RE (matrícula) — não deve ser modificado'
    if tipo == ERRO_CRITICO:
        return 'Coluna não deve ser modificada pelo pipeline — alteração inesperada'
    if tipo == ERRO_ESTRUTURAL:
        return 'Diferença estrutural (linhas ou abas)'
    return ''


def _selecionar_aba(wb):
    if NOME_ABA in wb.sheetnames:
        return wb[NOME_ABA]
    return wb['Frequência']


def comparar_arquivos(caminho_orig: str, caminho_proc: str) -> dict:
    erros = []
    estrutural = []
    resumo = Counter()

    wb_orig = openpyxl.load_workbook(caminho_orig, read_only=True, data_only=True)
    wb_proc = openpyxl.load_workbook(caminho_proc, read_only=True, data_only=True)

    sheets_orig = set(wb_orig.sheetnames)
    sheets_proc = set(wb_proc.sheetnames)

    if sheets_orig != sheets_proc:
        apenas_orig = sorted(sheets_orig - sheets_proc)
        apenas_proc = sorted(sheets_proc - sheets_orig)
        msg = []
        if apenas_orig:
            msg.append(f"Apenas no original: {apenas_orig}")
        if apenas_proc:
            msg.append(f"Apenas no processado: {apenas_proc}")
        estrutural.append({
            'tipo': ERRO_ESTRUTURAL,
            'descricao': 'Abas diferem. ' + '; '.join(msg),
        })
        resumo[ERRO_ESTRUTURAL] += 1

    for label, wb in [('original', wb_orig), ('processado', wb_proc)]:
        if NOME_ABA not in wb.sheetnames:
            acento = 'Frequência'
            if acento in wb.sheetnames:
                estrutural.append({
                    'tipo': ERRO_ESTRUTURAL,
                    'descricao': (
                        f"Aba '{NOME_ABA}' não encontrada no {label}; "
                        f"'{acento}' encontrada. Usando '{acento}'."
                    ),
                })
            else:
                estrutural.append({
                    'tipo': ERRO_ESTRUTURAL,
                    'descricao': f"Aba '{NOME_ABA}' ausente no arquivo {label}.",
                })
                wb_orig.close()
                wb_proc.close()
                return {
                    'erros': erros, 'resumo': resumo, 'estrutural': estrutural,
                    'total_linhas_orig': 0, 'total_linhas_proc': 0,
                    'total_linhas_comparadas': 0,
                }

    sh_orig = _selecionar_aba(wb_orig)
    sh_proc = _selecionar_aba(wb_proc)

    total_orig = 0
    total_proc = 0
    total_comparadas = 0
    _SENTINEL = object()

    iter_orig = sh_orig.iter_rows(values_only=True)
    iter_proc = sh_proc.iter_rows(values_only=True)

    while True:
        r1 = next(iter_orig, _SENTINEL)
        r2 = next(iter_proc, _SENTINEL)

        if r1 is _SENTINEL and r2 is _SENTINEL:
            break

        if r1 is not _SENTINEL:
            total_orig += 1
        if r2 is not _SENTINEL:
            total_proc += 1

        if r1 is _SENTINEL:
            erros.append({
                'linha': total_proc,
                'tipo': ERRO_ESTRUTURAL,
                'coluna_idx': -1,
                'coluna_nome': 'N/A',
                'valor_original': '(linha ausente)',
                'valor_processado': str(r2)[:200],
                're': '',
                'nome': '',
                'data': '',
                'contexto': 'Linha extra no arquivo processado',
            })
            resumo[ERRO_ESTRUTURAL] += 1
            continue

        if r2 is _SENTINEL:
            erros.append({
                'linha': total_orig,
                'tipo': ERRO_ESTRUTURAL,
                'coluna_idx': -1,
                'coluna_nome': 'N/A',
                'valor_original': str(r1)[:200],
                'valor_processado': '(linha ausente)',
                're': '',
                'nome': '',
                'data': '',
                'contexto': 'Linha ausente no arquivo processado',
            })
            resumo[ERRO_ESTRUTURAL] += 1
            continue

        total_comparadas += 1
        row_num = total_orig

        re_val   = r1[COL_RE]   if len(r1) > COL_RE   else None
        data_val = r1[COL_DATA] if len(r1) > COL_DATA else None
        nome_val = r1[COL_NOME] if len(r1) > COL_NOME else None
        re_str   = str(re_val).strip()   if re_val   is not None else ''
        data_str = normalizar_datetime(data_val)
        nome_str = str(nome_val).strip() if nome_val is not None else ''

        orig_18 = r1[COL_DESCONTOS] if len(r1) > COL_DESCONTOS else None
        proc_18 = r2[COL_DESCONTOS] if len(r2) > COL_DESCONTOS else None
        col18_mudou = (normalizar_time(orig_18) != normalizar_time(proc_18))

        n_cols = max(len(r1), len(r2), 26)
        for c in range(n_cols):
            v1 = r1[c] if c < len(r1) else None
            v2 = r2[c] if c < len(r2) else None

            n1 = normalizar_valor(v1, c)
            n2 = normalizar_valor(v2, c)

            if n1 == n2:
                continue

            tipo = classificar_erro(c, col18_mudou)
            col_nome = NOMES_COLUNAS.get(c, f'Col_{c}')

            erros.append({
                'linha': row_num,
                'tipo': tipo,
                'coluna_idx': c,
                'coluna_nome': col_nome,
                'valor_original': str(v1)[:300] if v1 is not None else '(vazio)',
                'valor_processado': str(v2)[:300] if v2 is not None else '(vazio)',
                're': re_str,
                'nome': nome_str,
                'data': data_str,
                'contexto': contexto_erro(tipo, c, col18_mudou),
            })
            resumo[tipo] += 1

    wb_orig.close()
    wb_proc.close()

    if total_orig != total_proc:
        estrutural.append({
            'tipo': ERRO_ESTRUTURAL,
            'descricao': (
                f"Número de linhas difere: original={total_orig}, "
                f"processado={total_proc}."
            ),
        })
        resumo[ERRO_ESTRUTURAL] += 1

    return {
        'erros': erros,
        'resumo': resumo,
        'estrutural': estrutural,
        'total_linhas_orig': total_orig,
        'total_linhas_proc': total_proc,
        'total_linhas_comparadas': total_comparadas,
    }
