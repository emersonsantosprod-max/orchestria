#!/usr/bin/env python3
"""
app.cli.validar_hr — Valida horas trabalhadas (col 19) na Medição.
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

from app.cli.validar_consist import COL_RE, COL_DATA, COL_HR_TRAB
from app.validar_horas import validar, gerar_relatorio, _salvar_relatorio

NOME_ABA = 'Frequencia'
DEFAULT_MEDICAO = 'data/entrada/medicao_base.xlsx'


def _get_sheet(wb):
    if NOME_ABA in wb.sheetnames:
        return wb[NOME_ABA]
    return wb['Frequência']


def _normalizar_data(valor) -> str:
    if valor is None:
        return ''
    if hasattr(valor, 'strftime'):
        return valor.strftime('%d/%m/%Y')
    return str(valor).strip()


def _ler_medicao(path: Path) -> tuple[list[dict], int]:
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        ws = _get_sheet(wb)
        registros: list[dict] = []
        n = 0
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            if row is None or all(c is None for c in row):
                continue
            n += 1
            mat_val = row[COL_RE]      if len(row) > COL_RE      else None
            dat_val = row[COL_DATA]    if len(row) > COL_DATA    else None
            hr_val  = row[COL_HR_TRAB] if len(row) > COL_HR_TRAB else None

            matricula = str(mat_val).strip() if mat_val is not None else ''
            data_str  = _normalizar_data(dat_val)
            hr_trab   = float(hr_val) if hr_val is not None else None

            registros.append({
                'matricula': matricula,
                'data': data_str,
                'hr_trabalhadas': hr_trab,
            })
    finally:
        wb.close()
    return registros, n


def cmd_validar(path: Path) -> int:
    if not path.exists():
        print(f"[ERRO] Arquivo não encontrado: {path}", file=sys.stderr)
        return 1

    registros, n_linhas = _ler_medicao(path)
    inconsistencias = validar(registros)

    conteudo = gerar_relatorio(inconsistencias, str(path.resolve()), n_linhas)
    caminho = _salvar_relatorio(conteudo)
    print(f"Relatório gravado em: {caminho}")
    print(f"Total de inconsistências: {len(inconsistencias)}")
    return 1 if inconsistencias else 0


def build_parser(parser: argparse.ArgumentParser) -> argparse.ArgumentParser:
    parser.add_argument(
        '--medicao',
        default=DEFAULT_MEDICAO,
        metavar='CAMINHO',
        help=f'Planilha de medição (default: {DEFAULT_MEDICAO})',
    )
    return parser


def main(argv=None) -> int:
    parser = build_parser(argparse.ArgumentParser(
        description='Valida horas trabalhadas (Hr Trabalhadas, col 19) na Medição.',
    ))
    args = parser.parse_args(argv)
    return cmd_validar(Path(args.medicao))


if __name__ == '__main__':
    sys.exit(main())
