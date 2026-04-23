"""
distribuicao_contratual.py — Normalização da distribuição contratual do BD.

Transforma o arquivo wide-format ("Dstribuição Contratual do BD - 2026.xlsx")
para o schema long-format: funcao | md_cobranca | area | quantidade.

Pipeline linear (STEP 1–5):
  1. Carrega e classifica headers (single-pass, read_only=True)
  2. Expande linhas em registros normalizados
  3. Agrega duplicatas de SIGLA
  4. Valida consistência interna
  5. Exporta xlsx normalizado
"""

from __future__ import annotations

import os
from collections import defaultdict

import openpyxl

ERRO_SIGLA               = 'ERRO_SIGLA'
ERRO_TOTAL               = 'ERRO_TOTAL'
AVISO_DECIMAL            = 'AVISO_DECIMAL'
AVISO_SIGLA_DUPLICADA    = 'AVISO_SIGLA_DUPLICADA'
AVISO_COLUNA_DESCONHECIDA = 'AVISO_COLUNA_DESCONHECIDA'
AVISO_COLUNA_DUPLICADA   = 'AVISO_COLUNA_DUPLICADA'
AVISO_VALOR_NAO_NUMERICO = 'AVISO_VALOR_NAO_NUMERICO'
AVISO_DISCREPANCIA_ATUAL = 'AVISO_DISCREPANCIA_ATUAL'
AVISO_HEADER_DUPLICADO   = 'AVISO_HEADER_DUPLICADO'

_SKIP_HEADERS = frozenset(
    ('SIGLA', 'TP MO', 'ÁREA', 'FUNÇÃO', 'Atual', 'OBSERVAÇÕES', 'INSPEÇÃO')
)

_AREA_MAP = {'PE1': 'PE-1', 'PE2': 'PE-2', 'PE3': 'PE-3'}


def normalize_area(raw: str) -> str:
    return _AREA_MAP.get(raw, raw)


def parse_distribuicao_cols(
    headers: tuple,
) -> tuple[dict[int, tuple[str, str | None]], list[dict]]:
    """
    Classifica cada coluna como distribuição ou skip com base no header.

    Returns:
      col_map:  {col_index: (md_cobranca, area)}
      warnings: list of inconsistency dicts
    """
    col_map: dict[int, tuple[str, str | None]] = {}
    warnings: list[dict] = []
    seen_keys: dict[tuple, int] = {}

    for i, h in enumerate(headers):
        if h is None:
            continue
        h_str = str(h).strip()
        if not h_str or h_str in _SKIP_HEADERS:
            continue

        if h_str == 'CENTRAL':
            key: tuple[str, str | None] = ('CENTRAL', None)
        elif h_str == 'ADM-B':
            key = ('ADM-B', None)
        elif h_str == 'ANALITICA':
            key = ('BREAKDOWN', 'ANALITICA')
        elif h_str.startswith('BREAKDOWN '):
            key = ('BREAKDOWN', normalize_area(h_str[len('BREAKDOWN '):].strip()))
        elif h_str.startswith('HD '):
            key = ('HD', normalize_area(h_str[len('HD '):].strip()))
        elif h_str.startswith('CV '):
            key = ('CV', normalize_area(h_str[len('CV '):].strip()))
        else:
            warnings.append({
                'tipo': AVISO_COLUNA_DESCONHECIDA,
                'coluna': h_str,
                'col_index': i,
                'erro': f"Header não reconhecido: '{h_str}'",
            })
            continue

        if key in seen_keys:
            warnings.append({
                'tipo': AVISO_COLUNA_DUPLICADA,
                'coluna': h_str,
                'col_index': i,
                'erro': (
                    f"Chave ({key[0]}, {key[1]}) já mapeada na col {seen_keys[key]}; "
                    f"quantidades serão somadas"
                ),
            })
        else:
            seen_keys[key] = i

        col_map[i] = key

    # Invariant: CENTRAL e ADM-B nunca têm area
    for (md, area) in col_map.values():
        if md in ('CENTRAL', 'ADM-B') and area is not None:
            raise ValueError(
                f"ERRO_CRITICO: coluna '{md}' mapeada com area='{area}' — invariante violado"
            )

    return col_map, warnings


def carregar_e_normalizar(
    path: str,
) -> tuple[list[dict], dict[str, float], dict[str, float], list[dict]]:
    """
    Encapsula STEP 1 (header), STEP 2 (row expansion), STEP 3 (aggregation).

    Returns:
      normalized   — list[{funcao, md_cobranca, area, quantidade}]
      raw_sums     — {sigla: soma das colunas classificadas na linha raw}
      atual        — {sigla: valor da coluna Atual}
      warnings     — todas as inconsistências dos STEPS 1-3
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    all_warnings: list[dict] = []
    col_map: dict[int, tuple[str, str | None]] = {}
    sigla_col: int | None = None
    funcao_col: int | None = None
    atual_col: int | None = None
    header_found = False
    header_count = 0

    expanded_records: list[tuple] = []
    expansion_warnings: list[dict] = []
    raw_sums: dict[str, float] = defaultdict(float)
    atual: dict[str, float] = {}
    sigla_to_funcao: dict[str, list[str]] = defaultdict(list)

    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if any(c == 'SIGLA' for c in row):
                header_count += 1
                if header_count == 1:
                    header_found = True
                    sigla_col = next(j for j, c in enumerate(row) if c == 'SIGLA')
                    funcao_col = next(
                        (j for j, c in enumerate(row) if c == 'FUNÇÃO'), None
                    )
                    atual_col = next(
                        (j for j, c in enumerate(row) if c == 'Atual'), None
                    )
                    col_map, parse_warnings = parse_distribuicao_cols(row)
                    all_warnings.extend(parse_warnings)
                elif header_count == 2:
                    all_warnings.append({
                        'tipo': AVISO_HEADER_DUPLICADO,
                        'erro': 'Mais de uma linha contém a célula "SIGLA"; usando a primeira',
                    })
            continue

        if sigla_col is not None and sigla_col < len(row) and row[sigla_col] == 'SIGLA':
            continue

        sigla_raw = row[sigla_col] if sigla_col < len(row) else None
        funcao_val = str(
            row[funcao_col] if funcao_col is not None and funcao_col < len(row) else ''
        ) or ''

        if not sigla_raw or (isinstance(sigla_raw, str) and not sigla_raw.strip()):
            has_qty = any(
                col_idx < len(row)
                and isinstance(row[col_idx], (int, float))
                and row[col_idx] not in (None, 0, 0.0)
                for col_idx in col_map
            )
            if has_qty:
                expansion_warnings.append({
                    'tipo': ERRO_SIGLA,
                    'funcao': funcao_val,
                    'erro': 'Linha com distribuição não-zero mas SIGLA ausente',
                })
            continue

        sigla = str(sigla_raw).strip()
        sigla_to_funcao[sigla].append(funcao_val)

        row_raw_sum = 0.0
        for col_idx, (md_cobranca, area) in col_map.items():
            if col_idx >= len(row):
                continue
            qty = row[col_idx]
            if qty is None or qty == 0:
                continue
            if not isinstance(qty, (int, float)):
                expansion_warnings.append({
                    'tipo': AVISO_VALOR_NAO_NUMERICO,
                    'funcao': sigla,
                    'col_index': col_idx,
                    'valor': repr(qty),
                    'erro': f"Valor não-numérico na col {col_idx}: {qty!r}; célula ignorada",
                })
                continue
            expanded_records.append((sigla, md_cobranca, area, qty))
            row_raw_sum += qty
            if isinstance(qty, float) and qty != int(qty):
                expansion_warnings.append({
                    'tipo': AVISO_DECIMAL,
                    'funcao': sigla,
                    'md_cobranca': md_cobranca,
                    'area': area,
                    'quantidade': qty,
                    'erro': f"Quantidade fracionária: {qty}",
                })

        raw_sums[sigla] += row_raw_sum
        if atual_col is not None and atual_col < len(row):
            v = row[atual_col]
            if isinstance(v, (int, float)):
                atual[sigla] = float(v)

    wb.close()

    if not header_found:
        raise ValueError("Header row not found — no cell equals 'SIGLA'")

    agg: dict[tuple, float] = defaultdict(float)
    for (funcao, md_cobranca, area, qty) in expanded_records:
        agg[(funcao, md_cobranca, area)] += qty

    normalized = [
        {'funcao': f, 'md_cobranca': m, 'area': a, 'quantidade': q}
        for (f, m, a), q in agg.items()
    ]

    agg_warnings: list[dict] = []
    for sigla, funcoes in sigla_to_funcao.items():
        if len(funcoes) > 1:
            agg_warnings.append({
                'tipo': AVISO_SIGLA_DUPLICADA,
                'funcao': sigla,
                'funcoes_origem': funcoes,
                'erro': f"SIGLA '{sigla}' aparece em {len(funcoes)} linhas: {funcoes}",
            })

    combined = all_warnings + expansion_warnings + agg_warnings
    return normalized, dict(raw_sums), atual, combined


def validar_distribuicao_cobranca(
    normalized: list[dict],
    raw_sums: dict[str, float],
    atual: dict[str, float],
) -> list[dict]:
    """
    Validação self-consistent: sum(normalized rows) deve igualar raw_sums por SIGLA.
    Separadamente, compara com a coluna Atual para AVISO_DISCREPANCIA_ATUAL.
    """
    inconsistencias: list[dict] = []

    norm_sums: dict[str, float] = defaultdict(float)
    for r in normalized:
        norm_sums[r['funcao']] += r['quantidade']

    for sigla, raw_sum in raw_sums.items():
        norm_sum = norm_sums.get(sigla, 0.0)
        if round(norm_sum, 10) != round(raw_sum, 10):
            inconsistencias.append({
                'tipo': ERRO_TOTAL,
                'funcao': sigla,
                'raw_sum': raw_sum,
                'normalized_sum': norm_sum,
                'erro': f"Soma normalizada ({norm_sum}) ≠ soma raw ({raw_sum})",
            })

        if sigla in atual:
            atu = atual[sigla]
            if round(norm_sum, 10) != round(float(atu), 10):
                inconsistencias.append({
                    'tipo': AVISO_DISCREPANCIA_ATUAL,
                    'funcao': sigla,
                    'normalized_sum': norm_sum,
                    'atual': atu,
                    'erro': f"Soma normalizada ({norm_sum}) ≠ coluna Atual ({atu})",
                })

    return inconsistencias


def exportar_normalizado(records: list[dict], path: str) -> None:
    """Escreve registros normalizados em xlsx. area=None → célula em branco."""
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)

    def sort_key(r: dict) -> tuple:
        return (r['funcao'], r['md_cobranca'], r['area'] or '')

    sorted_records = sorted(records, key=sort_key)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['funcao', 'md_cobranca', 'area', 'quantidade'])
    for r in sorted_records:
        ws.append([r['funcao'], r['md_cobranca'], r['area'], r['quantidade']])

    wb.save(path)
    wb.close()
