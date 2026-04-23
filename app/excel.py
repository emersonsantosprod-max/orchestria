"""
excel.py — Infraestrutura Excel: carregamento, mapeamento de colunas, indexação,
aplicação de updates e escrita via ZIP patch (performance crítica ~4s vs ~42s).

Normalização e dataclasses de contrato vivem em app/core.py.
"""

import os
import unicodedata
import zipfile
from xml.etree import ElementTree as ET

import openpyxl

from app.core import (
    converter_desconto_para_minutos as _core_converter_desconto_para_minutos,
)
from app.core import (
    converter_minutos_para_hhmmss as _core_converter_minutos_para_hhmmss,
)
from app.core import (
    deduplicar_observacao,
    inconsistencia,
)
from app.core import (
    normalizar_data as _core_normalizar_data,
)
from app.core import (
    normalizar_matricula as _core_normalizar_matricula,
)
from app.core import (
    parse_data_obj as _core_parse_data_obj,
)

_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
_NS_R = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
_NS_REL = 'http://schemas.openxmlformats.org/package/2006/relationships'

ET.register_namespace('', _NS)
ET.register_namespace('r', _NS_R)
ET.register_namespace('mc', 'http://schemas.openxmlformats.org/markup-compatibility/2006')
ET.register_namespace('x14ac', 'http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac')
ET.register_namespace('xr',  'http://schemas.microsoft.com/office/spreadsheetml/2014/revision')
ET.register_namespace('xr2', 'http://schemas.microsoft.com/office/spreadsheetml/2015/revision2')
ET.register_namespace('xr3', 'http://schemas.microsoft.com/office/spreadsheetml/2016/revision3')


def _col_letter(n: int) -> str:
    s = ''
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _cell_addr(row: int, col: int) -> str:
    return f"{_col_letter(col)}{row}"


def _normalizar_matricula(valor) -> str:
    return _core_normalizar_matricula(valor)


def _normalizar_data(valor) -> str:
    return _core_normalizar_data(valor)


def _parse_data_obj(valor):
    return _core_parse_data_obj(valor)


def _converter_desconto_para_minutos(valor) -> int:
    return _core_converter_desconto_para_minutos(valor)


def _converter_minutos_para_hhmmss(minutos: int) -> str:
    return _core_converter_minutos_para_hhmmss(minutos)


def carregar_planilha(caminho: str, read_only: bool = False, data_only: bool = False):
    if not os.path.exists(caminho):
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")

    wb = openpyxl.load_workbook(caminho, read_only=read_only, data_only=data_only)
    nome_aba = 'Frequência' if 'Frequência' in wb.sheetnames else 'Frequencia'
    if nome_aba not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Aba 'Frequência' não encontrada em: {caminho}")

    return wb, wb[nome_aba]


_PALAVRAS_PROIBIDAS_DESCONTO = {'total', 'apoio', 'dif', 'descanso'}


def _normalizar_header(s: str) -> str:
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return ' '.join(s.strip().lower().split())


def _match_coluna(header_norm: str, aliases: list, chave: str) -> bool:
    if header_norm in aliases:
        return True
    if chave in ('matricula', 'tag'):
        return False

    tokens = set(header_norm.split())

    for a in aliases:
        if set(a.split()).issubset(tokens):
            if chave == 'desconto' and (_PALAVRAS_PROIBIDAS_DESCONTO & tokens):
                continue
            return True

    if chave in ('desconto', 'observacao'):
        padded = f" {header_norm} "
        for a in aliases:
            if f" {a} " in padded:
                if chave == 'desconto' and (_PALAVRAS_PROIBIDAS_DESCONTO & tokens):
                    continue
                return True
    return False


_OBRIGATORIAS = ('data', 'matricula', 'desconto', 'observacao')


def mapear_colunas(sheet) -> dict:
    alvos = {
        'data':        ['data'],
        'matricula':   ['re', 'matricula'],
        'desconto':    ['descontos'],
        'observacao':  ['observacao', 'observacoes'],
        'situacao':     ['situacao'],
        'md_cobranca':  ['md cobranca'],
        'sg_funcao':    ['sg funcao'],
        'tag':          ['tag'],
        'pct_cobranca': ['% cobranca', 'pct cobranca'],
    }

    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=5, values_only=False), start=1
    ):
        colisoes = {}
        for cell in row:
            if not (cell.value and isinstance(cell.value, str)):
                continue
            header_norm = _normalizar_header(cell.value)
            for chave, aliases in alvos.items():
                if _match_coluna(header_norm, aliases, chave):
                    colisoes.setdefault(chave, []).append((cell.column, cell.value))

        ambig = {k: v for k, v in colisoes.items() if len(v) > 1}
        if ambig:
            detalhes = '; '.join(
                f"{k}: {[h for _, h in v]}" for k, v in ambig.items()
            )
            raise ValueError(f"[MAPEAMENTO] Ambiguidade linha {row_idx}: {detalhes}")

        mapeamento = {k: v[0][0] - 1 for k, v in colisoes.items()}

        if all(k in mapeamento for k in _OBRIGATORIAS):
            mapeamento['_header_row'] = row_idx
            return mapeamento

    raise ValueError(
        "Colunas 'Data', 'RE', 'Descontos' e 'Observação' não encontradas "
        "sem ambiguidade nas primeiras 5 linhas da planilha."
    )


def indexar_e_ler_dados(sheet, col_map: dict) -> tuple:
    """
    Retorna (sempre tupla de 7):
      index                : {(matricula, data_str): [row_1based, ...]}
      obs_existentes       : {(matricula, data_str): str}
      descontos_existentes : {(matricula, data_str): int_minutos}
      md_cobranca_por_chave: {(matricula, data_str): str_UPPER}
      sg_funcao_por_chave  : {(matricula, data_str): str_UPPER}
      medicao_por_matricula: {matricula: [(date_obj, data_str, [rows]), ...]}
      medicao_records      : list[dict] com (data, sg_funcao, md_cobranca,
                             pct_cobranca) por linha — usado por
                             validar_distribuicao.validar_para_dominio sem
                             segunda leitura. Requer col_map com sg_funcao,
                             md_cobranca e pct_cobranca; caso contrário list vazia.
    """
    index = {}
    obs_existentes = {}
    descontos_existentes = {}
    md_cobranca_por_chave = {}
    sg_funcao_por_chave = {}
    medicao_records: list[dict] = []

    col_re   = col_map['matricula']
    col_dt   = col_map['data']
    col_obs  = col_map['observacao']
    col_desc = col_map['desconto']
    col_mdc  = col_map.get('md_cobranca')
    col_sgf  = col_map.get('sg_funcao')
    col_pct  = col_map.get('pct_cobranca')
    hrow     = col_map['_header_row']

    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_idx <= hrow:
            continue

        re_val   = row[col_re]
        data_val = row[col_dt]
        if re_val is None or data_val is None:
            continue

        matricula = _core_normalizar_matricula(re_val)
        data_str  = _core_normalizar_data(data_val)
        chave     = (matricula, data_str)

        if chave not in index:
            index[chave] = []
        index[chave].append(row_idx)
        obs_existentes[chave] = str(row[col_obs]).strip() if row[col_obs] else ''
        descontos_existentes[chave] = _core_converter_desconto_para_minutos(row[col_desc])

        if col_mdc is not None and row[col_mdc] is not None:
            md_cobranca_por_chave[chave] = str(row[col_mdc]).strip().upper()
        if col_sgf is not None and row[col_sgf] is not None:
            sg_funcao_por_chave[chave] = str(row[col_sgf]).strip().upper()

        if col_sgf is not None and col_mdc is not None and col_pct is not None:
            sg = row[col_sgf]
            mdc = row[col_mdc]
            pct = row[col_pct]
            if sg is not None and mdc is not None:
                try:
                    pct_f = float(pct) if pct is not None else 0.0
                except (TypeError, ValueError):
                    pct_f = 0.0
                if pct_f > 1.0:
                    pct_f = pct_f / 100
                medicao_records.append({
                    'data':         data_str,
                    'sg_funcao':    str(sg).strip().upper(),
                    'md_cobranca':  str(mdc).strip().upper(),
                    'pct_cobranca': pct_f,
                })

    medicao_por_matricula = {}
    for (matricula, data_str), rows in index.items():
        data_obj = _core_parse_data_obj(data_str)
        if data_obj is None:
            continue
        medicao_por_matricula.setdefault(matricula, []).append(
            (data_obj, data_str, rows)
        )

    return (
        index,
        obs_existentes,
        descontos_existentes,
        md_cobranca_por_chave,
        sg_funcao_por_chave,
        medicao_por_matricula,
        medicao_records,
    )


def aplicar_updates(
    updates: list,
    col_map: dict,
    index: dict,
    obs_existentes: dict = None,
    descontos_existentes: dict = None,
) -> tuple:
    """
    Aplica a lista unificada de Update e retorna (patches, inconsistencias).

    Ordem: itera em ordem recebida (caller é responsável por colocar
    treinamento antes de férias para que a sobrescrita de observação vença).

    Regras:
      - row=None ⇒ resolve via index em todas as rows de (matricula, data);
        ausente → 'matrícula não encontrada' ou 'data não encontrada'.
      - row>0   ⇒ aplica diretamente; row<=0 ou não-int → 'row inválido'.
      - sobrescrever_obs=True  ⇒ escreve update.observacao na célula.
      - sobrescrever_obs=False ⇒ dedup com obs_existentes via core.deduplicar_observacao.
      - desconto_min>0 ⇒ soma com desconto existente (minutos) e formata HH:MM.
      - situacao ⇒ escrita apenas se coluna mapeada.

    patches: {(row_1based, col_1based): str_value}
    inconsistencias: list[Inconsistencia] (origem='writer')
    """
    if obs_existentes is None:
        obs_existentes = {}
    if descontos_existentes is None:
        descontos_existentes = {}

    max_row_known = max((r for rows in index.values() for r in rows), default=0)

    patches = {}
    inconsistencias = []

    col_obs_1  = col_map['observacao'] + 1
    col_desc_1 = col_map['desconto']   + 1
    col_sit_1  = (col_map['situacao'] + 1) if 'situacao' in col_map else None
    col_tag_1  = (col_map['tag'] + 1) if 'tag' in col_map else None

    matriculas_existentes = {m for m, _ in index}

    for upd in updates:
        mat = upd.matricula
        data = upd.data

        if upd.row is not None:
            if not isinstance(upd.row, int) or upd.row < 1 or (max_row_known > 0 and upd.row > max_row_known):
                inconsistencias.append(inconsistencia(
                    'writer', matricula=mat, data=data, erro='row inválido',
                ))
                continue
            target_rows = [upd.row]
        else:
            rows = index.get((mat, data))
            if not rows:
                erro = (
                    'data não encontrada'
                    if mat in matriculas_existentes
                    else 'matrícula não encontrada'
                )
                inconsistencias.append(inconsistencia(
                    'writer', matricula=mat, data=data, erro=erro,
                ))
                continue
            if isinstance(rows, int):
                rows = [rows]
            target_rows = rows

        obs_final = None
        if upd.observacao is not None:
            if upd.sobrescrever_obs:
                obs_final = upd.observacao
            else:
                novas = [p.strip() for p in upd.observacao.split(';') if p.strip()]
                existente = obs_existentes.get((mat, data), '')
                obs_final = deduplicar_observacao(existente, novas)

        # Desconto: soma aritmética com existente (mecânica, não decisão de negócio)
        desconto_str = None
        if upd.desconto_min is not None and upd.desconto_min > 0:
            existente_min = descontos_existentes.get((mat, data), 0)
            total_min = upd.desconto_min + existente_min
            desconto_str = _core_converter_minutos_para_hhmmss(total_min)

        sit = upd.situacao

        for row_idx in target_rows:
            if obs_final is not None:
                patches[(row_idx, col_obs_1)] = obs_final
            if desconto_str is not None:
                patches[(row_idx, col_desc_1)] = desconto_str
            if sit and col_sit_1 is not None:
                patches[(row_idx, col_sit_1)] = sit
            if upd.tipo == 'atestado' and col_tag_1 is not None:
                patches[(row_idx, col_tag_1)] = 'ATESTADO'

    return patches, inconsistencias


# Escrita via ZIP patch — NÃO substituir por openpyxl write mode (CLAUDE.md CRITICAL: 4s → 42s).
def _set_string_cell(c_el, value: str):
    for child in list(c_el):
        c_el.remove(child)

    if value:
        c_el.set('t', 'inlineStr')
        is_el = ET.SubElement(c_el, f'{{{_NS}}}is')
        t_el  = ET.SubElement(is_el, f'{{{_NS}}}t')
        t_el.text = value
    else:
        c_el.attrib.pop('t', None)


def salvar_via_zip(
    src: str,
    dst: str,
    patches: dict,
    nome_aba: str = 'Frequencia',
):
    addr_patches = {
        _cell_addr(row, col): val
        for (row, col), val in patches.items()
    }
    addr_set = set(addr_patches)

    with zipfile.ZipFile(src) as z:
        wb_xml   = ET.fromstring(z.read('xl/workbook.xml'))
        rels_xml = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))

        rid_to_file = {}
        for rel in rels_xml.iter(f'{{{_NS_REL}}}Relationship'):
            target = rel.get('Target')
            if not target:
                continue
            if target.startswith('/'):
                target = target[1:]
            if not target.startswith('xl/'):
                target = 'xl/' + target
            rid_to_file[rel.get('Id')] = target

        nome_norm = nome_aba.lower().replace('ê', 'e')
        sheet_path = None
        for sheet in wb_xml.iter(f'{{{_NS}}}sheet'):
            if sheet.get('name', '').lower().replace('ê', 'e') == nome_norm:
                rid = sheet.get(f'{{{_NS_R}}}id')
                sheet_path = rid_to_file[rid]
                break

        if not sheet_path:
            raise ValueError(f"Aba '{nome_aba}' não encontrada no workbook.xml")

        xml_bytes = z.read(sheet_path)
        outros = [
            (info, z.read(info.filename))
            for info in z.infolist()
            if info.filename != sheet_path
            and info.filename != 'xl/calcChain.xml'
        ]

    root = ET.fromstring(xml_bytes)
    del xml_bytes

    for c in root.iter(f'{{{_NS}}}c'):
        r = c.get('r')
        if r and r in addr_set:
            _set_string_cell(c, addr_patches[r])

    xml_str = ET.tostring(root, encoding='unicode')

    for prefix, uri in [
        ('xr2', 'http://schemas.microsoft.com/office/spreadsheetml/2015/revision2'),
        ('xr3', 'http://schemas.microsoft.com/office/spreadsheetml/2016/revision3'),
    ]:
        if f'xmlns:{prefix}=' not in xml_str:
            xml_str = xml_str.replace('<worksheet ', f'<worksheet xmlns:{prefix}="{uri}" ', 1)

    new_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        + xml_str
    ).encode('utf-8')
    del root

    diretorio = os.path.dirname(dst)
    if diretorio:
        os.makedirs(diretorio, exist_ok=True)

    try:
        with zipfile.ZipFile(dst, 'w', compression=zipfile.ZIP_DEFLATED) as z_new:
            for info, data in outros:
                z_new.writestr(info, data)
            z_new.writestr(sheet_path, new_xml)
    except PermissionError as err:
        raise PermissionError(
            f"Erro ao salvar: feche o arquivo '{os.path.basename(dst)}' antes de prosseguir."
        ) from err
