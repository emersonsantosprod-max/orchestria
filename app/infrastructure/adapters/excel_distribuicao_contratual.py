"""
Adapter Excel para a distribuição contratual.

Camada de I/O: abre o xlsx wide-format, localiza o header (linha que contém a
célula 'SIGLA'), retorna headers + linhas de dados como tuplas. Sem regras de
negócio: classificação de colunas, normalização de área e agregação vivem em
``app.domain.distribuicao_contratual``.

Falhas estruturais (workbook vazio, header ausente) levantam
:class:`DistribuicaoContratualMalformadaError`.
"""

from __future__ import annotations

import os
from pathlib import Path

import openpyxl

from app.domain.distribuicao_contratual import AVISO_HEADER_DUPLICADO


class DistribuicaoContratualMalformadaError(ValueError):
    """Estrutura inválida do xlsx contratual: ausência de header SIGLA."""


def _normalizar_celula(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        return v.strip()
    return v


def ler_xlsx_contratual(
    path: str | Path,
) -> tuple[tuple, list[tuple], list[dict]]:
    """Lê o xlsx contratual e devolve ``(headers, data_rows, header_warnings)``.

    - ``headers``: tupla com o conteúdo da primeira linha que contém a célula
      ``'SIGLA'``.
    - ``data_rows``: linhas seguintes (sem o header), cada uma como tupla.
    - ``header_warnings``: lista contendo ``AVISO_HEADER_DUPLICADO`` se mais
      de uma linha contiver a célula ``'SIGLA'``.

    Levanta :class:`DistribuicaoContratualMalformadaError` se nenhuma linha
    contiver a célula ``'SIGLA'``.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        all_rows = [
            tuple(_normalizar_celula(c) for c in row)
            for row in ws.iter_rows(values_only=True)
        ]
    finally:
        wb.close()

    header_idx = next(
        (i for i, row in enumerate(all_rows) if any(c == 'SIGLA' for c in row)),
        None,
    )
    if header_idx is None:
        raise DistribuicaoContratualMalformadaError(
            "Header não encontrado: nenhuma linha contém a célula 'SIGLA'"
        )

    headers = all_rows[header_idx]
    data_rows = all_rows[header_idx + 1:]

    extra_header_rows = sum(
        1 for row in data_rows if any(c == 'SIGLA' for c in row)
    )
    header_warnings: list[dict] = []
    if extra_header_rows > 0:
        header_warnings.append({
            'tipo': AVISO_HEADER_DUPLICADO,
            'erro': 'Mais de uma linha contém a célula "SIGLA"; usando a primeira',
        })

    return headers, data_rows, header_warnings


def escrever_xlsx_normalizado(records: list[dict], path: str | Path) -> None:
    """Persiste registros normalizados em xlsx, ordenando por (funcao, md, area)."""
    parent = os.path.dirname(os.path.abspath(str(path)))
    if parent:
        os.makedirs(parent, exist_ok=True)

    def _sort_key(r: dict) -> tuple:
        return (r['funcao'], r['md_cobranca'], r['area'] or '')

    sorted_records = sorted(records, key=_sort_key)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['funcao', 'md_cobranca', 'area', 'quantidade'])
    for r in sorted_records:
        ws.append([r['funcao'], r['md_cobranca'], r['area'], r['quantidade']])
    wb.save(path)
    wb.close()
