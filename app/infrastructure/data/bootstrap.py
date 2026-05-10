"""
bootstrap.py — orquestração xlsx → SQLite + leituras de alto nível.

Substitui o antigo `app.infrastructure.db` (deletado). Usa os Repositories
em `app/infrastructure/data/repositories/` para CRUD e adiciona apenas a
camada de orquestração (parse xlsx, transação, registro de arquivo).

Regra: bootstrap.py é o único lugar (junto à composition root) que abre
transação e faz commit. Os Repositories permanecem livres de commit.
"""

from __future__ import annotations

import logging
import sqlite3
from collections.abc import Iterator
from contextlib import closing, contextmanager
from pathlib import Path

import openpyxl

from app.domain.core import normalizar_data
from app.domain.errors import PlanilhaInvalidaError
from app.domain.reference_month import mes_referencia_unico
from app.infrastructure.data.registry import RegistryRepository
from app.infrastructure.data.repositories.distribuicao import DistribuicaoRepository
from app.infrastructure.data.repositories.ferias import FeriasRepository
from app.infrastructure.data.repositories.treinamentos import TreinamentosRepository

logger = logging.getLogger(__name__)

_HEADER_SCAN_ROWS = 20

_ALIASES_MEDICAO = {
    'data':         {'data'},
    'sg_funcao':    {'sg funcao', 'sg função'},
    'md_cobranca':  {'md cobranca', 'md cobrança'},
    'pct_cobranca': {'% cobrança', '% cobranca'},
}


def _normalizar_pct(value) -> float:
    if value is None:
        return 0.0
    v = float(value)
    return v / 100 if v > 1.0 else v


@contextmanager
def _abrir_aba_frequencia(
    path: str | Path,
) -> Iterator[tuple[object, dict[str, int], int]]:
    """Abre o workbook da Medição em modo leitura, resolve o cabeçalho da
    aba 'Frequencia' e produz `(worksheet, col_map, header_row_idx)`.

    `header_row_idx` é 1-based (compatível com `min_row` do openpyxl) e
    indica a linha onde o cabeçalho foi encontrado. Garante `wb.close()`
    no `__exit__`, mesmo em early-return ou exceção do consumidor — evita
    vazamento de file handle do openpyxl.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        nome = next(
            (n for n in wb.sheetnames if n.lower() in {'frequencia', 'frequência'}),
            None,
        )
        if nome is None:
            raise ValueError(
                "Aba 'Frequencia' não encontrada na planilha de medição."
            )
        ws = wb[nome]
        col_map: dict[str, int] | None = None
        header_row_idx: int = 0
        linhas_dados_scan = 0
        row_idx = 0
        for row in ws.iter_rows(values_only=True):
            row_idx += 1
            if all(cell is None for cell in row):
                continue
            linhas_dados_scan += 1
            if linhas_dados_scan > _HEADER_SCAN_ROWS:
                break
            candidate: dict[str, int] = {}
            for col_i, cell in enumerate(row):
                if cell is None:
                    continue
                label = str(cell).strip().lower()
                for field, aliases in _ALIASES_MEDICAO.items():
                    if label in aliases and field not in candidate:
                        candidate[field] = col_i
            if len(candidate) == len(_ALIASES_MEDICAO):
                col_map = candidate
                header_row_idx = row_idx
                break
        if col_map is None:
            raise ValueError(
                'Cabeçalho da Medição não encontrado nas primeiras '
                f'{_HEADER_SCAN_ROWS} linhas com dados. '
                'Colunas esperadas: data, sg funcao, md cobranca, % cobrança'
            )
        yield ws, col_map, header_row_idx
    finally:
        wb.close()


def _iter_linhas_dados(
    ws,
    col_map: dict[str, int],
    header_row_idx: int,
) -> Iterator[tuple[str, str, str, object]]:
    """Itera linhas de dados da aba Frequencia (a partir de header+1),
    normalizando data/funcao/md_cobranca.

    Pula linhas sem matrícula útil (sg_funcao + md_cobranca obrigatórios).
    Retorna `(data_str, sg_funcao_str, md_cobranca_str, pct_val_bruto)`;
    a normalização de pct fica a cargo do consumidor.
    """
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        n = len(row)
        data_val        = row[col_map['data']]         if col_map['data']         < n else None
        sg_funcao_val   = row[col_map['sg_funcao']]    if col_map['sg_funcao']    < n else None
        md_cobranca_val = row[col_map['md_cobranca']]  if col_map['md_cobranca']  < n else None
        pct_val         = row[col_map['pct_cobranca']] if col_map['pct_cobranca'] < n else None

        if data_val is None and sg_funcao_val is None:
            continue

        sg_funcao_str   = str(sg_funcao_val).strip().upper()   if sg_funcao_val   is not None else ''
        md_cobranca_str = str(md_cobranca_val).strip().upper() if md_cobranca_val is not None else ''

        if not sg_funcao_str or not md_cobranca_str:
            continue

        data_str = normalizar_data(data_val)
        yield data_str, sg_funcao_str, md_cobranca_str, pct_val


def ler_medicao_do_excel(path: str | Path) -> tuple[list[dict], list[str]]:
    """Lê a aba Frequencia do Excel da medição e retorna (registros, avisos).

    Não toca em SQLite. Cada registro: dict com chaves
    `data, sg_funcao, md_cobranca, pct_cobranca` (pct normalizado para 0-1).
    """
    registros: list[dict] = []
    tem_maior_1 = False
    tem_menor_igual_1 = False

    with _abrir_aba_frequencia(path) as (ws, col_map, header_row_idx):
        for data_str, sg_funcao_str, md_cobranca_str, pct_val in _iter_linhas_dados(ws, col_map, header_row_idx):
            if pct_val is not None:
                pv = float(pct_val)
                if pv > 1.0:
                    tem_maior_1 = True
                else:
                    tem_menor_igual_1 = True
            registros.append({
                'data': data_str,
                'sg_funcao': sg_funcao_str,
                'md_cobranca': md_cobranca_str,
                'pct_cobranca': _normalizar_pct(pct_val),
            })

    avisos: list[str] = []
    if tem_maior_1 and tem_menor_igual_1:
        avisos.append(
            'AVISO_ESCALA_INDEFINIDA: coluna % Cobrança contém valores em '
            'escalas mistas (alguns > 1.0, alguns ≤ 1.0). '
            'Normalização aplicada linha a linha.'
        )
    return registros, avisos


def obter_mes_referencia_medicao(path: str | Path) -> str:
    """Retorna 'YYYY-MM' do mês único da medição.

    Levanta PlanilhaInvalidaError se as datas atravessarem múltiplos
    meses ou se não houver datas válidas. Streaming: para na primeira
    divergência.
    """
    def _pares() -> Iterator[tuple[int, int] | None]:
        with _abrir_aba_frequencia(path) as (ws, col_map, header_row_idx):
            for data_str, _sg, _md, _pct in _iter_linhas_dados(ws, col_map, header_row_idx):
                if not data_str or len(data_str) != 10:
                    yield None
                    continue
                yield (int(data_str[6:10]), int(data_str[3:5]))

    ano, mes = mes_referencia_unico(_pares(), contexto="Medição")
    return f'{ano:04d}-{mes:02d}'


def obter_mes_referencia_medicao_lite(path: str | Path) -> str:
    """Extração lite no register-time: primeira data válida vence.

    Localiza apenas a coluna 'data' na aba Frequencia (não exige header
    completo de cobranca/funcao/pct). Não chama mes_referencia_unico
    (premissa: medição cobre sempre um único mês). Validação estrita
    roda em pipeline._mes_referencia no Execute como defesa em
    profundidade.
    """
    with closing(openpyxl.load_workbook(path, read_only=True, data_only=True)) as wb:
        nome = next(
            (n for n in wb.sheetnames if n.lower() in {'frequencia', 'frequência'}),
            None,
        )
        if nome is None:
            raise PlanilhaInvalidaError(
                "Aba 'Frequencia' não encontrada na planilha de medição."
            )
        ws = wb[nome]
        col_data: int | None = None
        header_row_idx = 0
        rows_scan = 0
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if all(cell is None for cell in row):
                continue
            rows_scan += 1
            if rows_scan > _HEADER_SCAN_ROWS:
                break
            for col_i, cell in enumerate(row):
                if cell is None:
                    continue
                if str(cell).strip().lower() == 'data':
                    col_data = col_i
                    header_row_idx = row_idx
                    break
            if col_data is not None:
                break
        if col_data is None:
            raise PlanilhaInvalidaError(
                "Coluna 'data' não encontrada na aba Frequencia."
            )
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if col_data >= len(row):
                continue
            v = row[col_data]
            if v is None:
                continue
            data_str = normalizar_data(v)
            if data_str and len(data_str) == 10:
                ano = int(data_str[6:10])
                mes = int(data_str[3:5])
                return f'{ano:04d}-{mes:02d}'
    raise PlanilhaInvalidaError('Medição sem datas válidas')


def registrar_bd(path: str | Path, conn: sqlite3.Connection) -> None:
    logger.info('registrar_bd: lendo %s', path)
    with closing(openpyxl.load_workbook(path, read_only=True, data_only=True)) as wb:
        ws = wb.active
        records: list[tuple] = []
        header_skipped = False
        for row in ws.iter_rows(values_only=True):
            if not header_skipped:
                header_skipped = True
                continue
            if len(row) < 4 or row[0] is None:
                continue
            funcao      = str(row[0]).strip().upper()
            md_cobranca = str(row[1]).strip().upper() if row[1] is not None else ''
            area        = str(row[2]).strip() if row[2] is not None else None
            quantidade  = float(row[3]) if row[3] is not None else 0.0
            records.append((funcao, md_cobranca, area, quantidade))
    logger.info('registrar_bd: %d linhas extraídas', len(records))

    DistribuicaoRepository(conn).salvar(records)
    RegistryRepository(conn).upsert('bd', str(path))
    conn.commit()


def registrar_medicao_arquivo(
    path: str | Path,
    conn: sqlite3.Connection,
) -> tuple[list[str], int]:
    """Valida o Excel da medição e registra o caminho durável em registro_arquivos.

    Não persiste o conteúdo em SQLite — leitura runtime via `obter_medicao(conn)`
    re-lê o Excel apontado por `registro_arquivos['medicao']`.

    Retorna `(avisos, total_registros)`. `path` deve apontar para arquivo
    durável e legível; rotas de registro validam via
    `validar_arquivo_referenciado` antes de chamar.
    """
    registros, avisos = ler_medicao_do_excel(path)
    RegistryRepository(conn).upsert('medicao', str(path))
    conn.commit()
    return avisos, len(registros)


def registrar_base_treinamentos(path: str | Path, conn: sqlite3.Connection) -> None:
    """Importa Base de Treinamentos para `catalogo_treinamentos`."""
    logger.info('registrar_base_treinamentos: lendo %s', path)
    records: list[tuple] = []
    with closing(openpyxl.load_workbook(path, read_only=True, data_only=True)) as wb:
        for row in wb.active.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            nome = str(row[0]).strip().upper()
            tipo_raw = str(row[1]).strip().lower() if row[1] else ''
            tipo = 'nao_remunerado' if ('não' in tipo_raw or 'nao' in tipo_raw) else 'remunerado'
            records.append((nome, tipo))
    logger.info('registrar_base_treinamentos: %d treinamentos extraídos', len(records))
    if not records:
        raise PlanilhaInvalidaError('Base de Treinamentos sem dados')

    TreinamentosRepository(conn).salvar(records)
    RegistryRepository(conn).upsert('treinamentos', str(path))
    conn.commit()


def registrar_cobranca(path: str | Path, conn: sqlite3.Connection) -> None:
    """Importa base_cobranca.xlsx em `regras_pagamento_ferias`.

    Convenção: `remunerado=0` sse md_cobranca == 'FÉRIAS S/ DESC' (uppercased),
    senão 1 — alinha com o domínio de férias.
    """
    logger.info('registrar_cobranca: lendo %s', path)
    records: list[tuple] = []
    with closing(openpyxl.load_workbook(path, read_only=True, data_only=True)) as wb:
        for row in wb.active.iter_rows(values_only=True):
            if row[0] is None:
                continue
            sg_funcao = str(row[0]).strip().upper()
            md_cobranca = (
                str(row[1]).strip().upper() if len(row) > 1 and row[1] is not None else ''
            )
            if not sg_funcao:
                continue
            remunerado = 0 if md_cobranca == 'FÉRIAS S/ DESC' else 1
            records.append((sg_funcao, md_cobranca, remunerado))
    logger.info('registrar_cobranca: %d linhas extraídas', len(records))
    if not records:
        raise PlanilhaInvalidaError('Base de Férias sem dados')

    FeriasRepository(conn).salvar(records)
    RegistryRepository(conn).upsert('cobranca', str(path))
    conn.commit()


def obter_bd(conn: sqlite3.Connection) -> list[dict]:
    return DistribuicaoRepository(conn).listar()


def obter_medicao(conn: sqlite3.Connection) -> list[dict]:
    """Retorna os registros da medição re-lendo o Excel registrado.

    Falha de forma terminal se: (a) nenhum arquivo registrado em
    `registro_arquivos['medicao']`; (b) caminho registrado não existe em disco
    (corrupção operacional, não estado vazio).
    """
    registro = RegistryRepository(conn).get('medicao')
    if registro is None:
        raise FileNotFoundError(
            'Nenhum arquivo de medição registrado. Faça upload via '
            '/api/config/medicao ou --registrar-medicao.'
        )
    caminho = registro['caminho']
    if not Path(caminho).exists():
        raise FileNotFoundError(
            f'Arquivo de medição registrado não encontrado: {caminho}'
        )
    return ler_medicao_do_excel(caminho)[0]


def obter_mes_referencia_relatorio_treinamento(path: str | Path) -> str:
    """Retorna 'YYYY-MM' do mês único do relatório mensal de Treinamentos.

    Lê coluna `data` (col 6) a partir da linha 3 (mesmo schema do
    `loaders.carregar_dados_treinamento`). Streaming + early-exit.
    Levanta PlanilhaInvalidaError em multi-mês ou planilha sem datas.

    Duplicação de parsing com o loader é intencional nesta camada — a
    unificação single-pass é objeto de roadmap separado.
    """
    def _pares() -> Iterator[tuple[int, int] | None]:
        with closing(
            openpyxl.load_workbook(path, read_only=True, data_only=True)
        ) as wb:
            for row in wb.active.iter_rows(min_row=3, values_only=True):
                if not row or len(row) <= 6:
                    continue
                d = row[6]
                if d is None:
                    yield None
                    continue
                ano = getattr(d, 'year', None)
                mes = getattr(d, 'month', None)
                if ano is None or mes is None:
                    yield None
                    continue
                yield (int(ano), int(mes))

    ano, mes = mes_referencia_unico(_pares(), contexto="Relatório de Treinamentos")
    return f'{ano:04d}-{mes:02d}'


def obter_medicao_atual(conn: sqlite3.Connection) -> dict | None:
    """Retorna o registro atual da medição em SQLite ({'caminho','importado_em'})
    ou None se não houver upload registrado.

    Distinto de `obter_medicao(conn)` — não relê o Excel; é o lookup
    barato usado por gating (e.g., validação de mês na rota de catálogo).
    """
    return RegistryRepository(conn).get('medicao')


def obter_cobranca(conn: sqlite3.Connection) -> dict[str, str]:
    return FeriasRepository(conn).obter_mapa()


def obter_tabela_treinamento(conn: sqlite3.Connection) -> dict[str, str]:
    return TreinamentosRepository(conn).obter()


def obter_registro_arquivos(conn: sqlite3.Connection) -> dict[str, dict]:
    return RegistryRepository(conn).get_all()


def popular_cobranca_se_vazio(conn: sqlite3.Connection, xlsx: Path | str | None = None) -> bool:
    """Bootstrap idempotente de regras_pagamento_ferias a partir de xlsx informado."""
    if FeriasRepository(conn).count() > 0:
        return False
    if RegistryRepository(conn).get('cobranca') is not None:
        return False
    if xlsx is None or not Path(xlsx).exists():
        return False
    try:
        registrar_cobranca(xlsx, conn)
    except Exception:
        conn.rollback()
        logger.exception('popular_cobranca_se_vazio: rollback após falha')
        raise
    return True
