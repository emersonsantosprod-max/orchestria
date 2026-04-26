"""
db.py — persistência leve em SQLite para BD de distribuição e Medição.

Responsabilidades:
  - Criar e manter o schema do banco local
  - Registrar (importar) BD normalizado e Medição do xlsx
  - Expor leituras simples para o módulo de validação

Regras:
  - Leitura de workbooks: read_only=True, data_only=True (CLAUDE.md)
  - Passada única em iter_rows — sem acesso aleatório por coordenada
  - Normalização de % Cobrança: value > 1.0 → /100; ≤ 1.0 → as-is
  - Caminho do arquivo SQLite é resolvido por app.paths.db_path(); nunca
    referenciar Path('data/automacao.db') diretamente neste módulo.

NOTE: registrar_medicao / obter_medicao servem apenas ao fluxo CLI/GUI de
Validação autônoma. pipeline.executar_pipeline() nunca usa essas funções —
medição é sempre ephemeral dentro do pipeline.
"""

from __future__ import annotations

import logging
import sqlite3
from datetime import datetime
from pathlib import Path

import openpyxl

from app.domain.core import normalizar_data
from app.infrastructure.paths import bundled_distribuicao_xlsx, bundled_treinamentos_xlsx, db_path

logger = logging.getLogger(__name__)

_HEADER_SCAN_ROWS = 20

_ALIASES_MEDICAO = {
    'data':         {'data'},
    'sg_funcao':    {'sg funcao', 'sg função'},
    'md_cobranca':  {'md cobranca', 'md cobrança'},
    'pct_cobranca': {'% cobrança', '% cobranca'},
}

_SCHEMA = """
CREATE TABLE IF NOT EXISTS bd_distribuicao (
    funcao      TEXT NOT NULL,
    md_cobranca TEXT NOT NULL,
    area        TEXT,
    quantidade  REAL NOT NULL
);
CREATE TABLE IF NOT EXISTS medicao_frequencia (
    data         TEXT NOT NULL,
    sg_funcao    TEXT NOT NULL,
    md_cobranca  TEXT NOT NULL,
    pct_cobranca REAL NOT NULL
);
CREATE TABLE IF NOT EXISTS bd_treinamentos (
    nome TEXT NOT NULL,
    tipo TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS registro_arquivos (
    tipo         TEXT PRIMARY KEY,
    caminho      TEXT NOT NULL,
    importado_em TEXT NOT NULL
);
"""


def conectar(path: Path | str | None = None) -> sqlite3.Connection:
    """Abre conexão no caminho canônico (app.paths.db_path()).

    `path` pode ser sobrescrito (útil em testes com tmp_path ou ':memory:').

    Hardening: timeout=5s + WAL + busy_timeout=5000 garantem que travas
    pendentes de runs anteriores transformam um wait infinito em
    OperationalError dentro de ~5s — recuperável pela boundary GUI/CLI.
    """
    resolved = path if path is not None else db_path()
    logger.info('db.conectar: abrindo %s', resolved)
    if str(resolved) != ':memory:':
        Path(resolved).parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(resolved, timeout=5)
    conn.row_factory = sqlite3.Row
    if str(resolved) != ':memory:':
        conn.execute('PRAGMA journal_mode=WAL')
    conn.execute('PRAGMA busy_timeout=5000')
    conn.executescript(_SCHEMA)
    conn.commit()
    logger.info('db.conectar: conexão pronta (%s)', resolved)
    return conn


def _normalizar_pct(value) -> float:
    if value is None:
        return 0.0
    v = float(value)
    return v / 100 if v > 1.0 else v


def registrar_bd(path: str | Path, conn: sqlite3.Connection) -> None:
    logger.info('registrar_bd: lendo %s', path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    records = []
    header_skipped = False
    for row in ws.iter_rows(values_only=True):
        if not header_skipped:
            header_skipped = True
            continue
        if len(row) < 4 or row[0] is None:
            continue
        funcao      = str(row[0]).strip().upper()
        md_cobranca = str(row[1]).strip().upper() if row[1] is not None else ''
        area        = str(row[2]).strip() if row[2] is not None else None
        quantidade  = float(row[3]) if row[3] is not None else 0.0
        records.append((funcao, md_cobranca, area, quantidade))
    wb.close()
    logger.info('registrar_bd: %d linhas extraídas', len(records))

    conn.execute('DELETE FROM bd_distribuicao')
    conn.executemany(
        'INSERT INTO bd_distribuicao (funcao, md_cobranca, area, quantidade) VALUES (?,?,?,?)',
        records,
    )
    conn.execute(
        'INSERT OR REPLACE INTO registro_arquivos (tipo, caminho, importado_em) VALUES (?,?,?)',
        ('bd', str(path), datetime.now().isoformat()),
    )
    conn.commit()


def registrar_medicao(path: str | Path, conn: sqlite3.Connection) -> list[str]:
    """Import Medição Frequencia sheet into SQLite. Returns list of warnings."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['Frequencia']

    col_map: dict[str, int] | None = None
    records: list[tuple] = []
    tem_maior_1 = False
    tem_menor_igual_1 = False
    linhas_dados_scan = 0

    for row in ws.iter_rows(values_only=True):
        if col_map is None:
            if all(cell is None for cell in row):
                continue
            linhas_dados_scan += 1
            if linhas_dados_scan > _HEADER_SCAN_ROWS:
                break
            candidate: dict[str, int] = {}
            for col_i, cell in enumerate(row):
                if cell is None:
                    continue
                label = str(cell).strip().lower()
                for field, aliases in _ALIASES_MEDICAO.items():
                    if label in aliases and field not in candidate:
                        candidate[field] = col_i
            if len(candidate) == len(_ALIASES_MEDICAO):
                col_map = candidate
            continue

        n = len(row)
        data_val        = row[col_map['data']]         if col_map['data']         < n else None
        sg_funcao_val   = row[col_map['sg_funcao']]    if col_map['sg_funcao']    < n else None
        md_cobranca_val = row[col_map['md_cobranca']]  if col_map['md_cobranca']  < n else None
        pct_val         = row[col_map['pct_cobranca']] if col_map['pct_cobranca'] < n else None

        if data_val is None and sg_funcao_val is None:
            continue

        sg_funcao_str   = str(sg_funcao_val).strip().upper()   if sg_funcao_val   is not None else ''
        md_cobranca_str = str(md_cobranca_val).strip().upper() if md_cobranca_val is not None else ''

        if not sg_funcao_str or not md_cobranca_str:
            continue

        data_str = normalizar_data(data_val)

        if pct_val is not None:
            pv = float(pct_val)
            if pv > 1.0:
                tem_maior_1 = True
            else:
                tem_menor_igual_1 = True

        records.append((data_str, sg_funcao_str, md_cobranca_str, _normalizar_pct(pct_val)))

    wb.close()

    if col_map is None:
        raise ValueError(
            'Cabeçalho da Medição não encontrado nas primeiras '
            f'{_HEADER_SCAN_ROWS} linhas com dados. '
            'Colunas esperadas: data, sg funcao, md cobranca, % cobrança'
        )

    avisos: list[str] = []
    if tem_maior_1 and tem_menor_igual_1:
        avisos.append(
            'AVISO_ESCALA_INDEFINIDA: coluna % Cobrança contém valores em '
            'escalas mistas (alguns > 1.0, alguns ≤ 1.0). '
            'Normalização aplicada linha a linha.'
        )

    conn.execute('DELETE FROM medicao_frequencia')
    conn.executemany(
        'INSERT INTO medicao_frequencia (data, sg_funcao, md_cobranca, pct_cobranca) VALUES (?,?,?,?)',
        records,
    )
    conn.execute(
        'INSERT OR REPLACE INTO registro_arquivos (tipo, caminho, importado_em) VALUES (?,?,?)',
        ('medicao', str(path), datetime.now().isoformat()),
    )
    conn.commit()
    return avisos


def popular_bd_se_vazio(conn: sqlite3.Connection) -> bool:
    """Bootstrap idempotente da tabela bd_distribuicao a partir do xlsx empacotado.

    Popula se e somente se:
      - bd_distribuicao está vazia  E
      - registro_arquivos não tem entrada para tipo='bd'

    Executa em transação única; rollback em falha. Retorna True se populou,
    False se já havia dados ou se o xlsx empacotado não estiver disponível.
    """
    logger.info('popular_bd_se_vazio: verificando estado atual')
    row_count = conn.execute('SELECT COUNT(*) FROM bd_distribuicao').fetchone()[0]
    reg = conn.execute(
        "SELECT 1 FROM registro_arquivos WHERE tipo='bd' LIMIT 1"
    ).fetchone()
    if row_count > 0 or reg is not None:
        logger.info('popular_bd_se_vazio: já populado, no-op')
        return False

    xlsx = bundled_distribuicao_xlsx()
    if not xlsx.exists():
        logger.info('popular_bd_se_vazio: xlsx empacotado ausente em %s', xlsx)
        return False

    logger.info('popular_bd_se_vazio: bootstrap a partir de %s', xlsx)
    try:
        conn.execute('BEGIN')
        registrar_bd(xlsx, conn)
    except Exception:
        conn.rollback()
        logger.exception('popular_bd_se_vazio: rollback após falha')
        raise
    logger.info('popular_bd_se_vazio: bootstrap concluído')
    return True


def popular_treinamentos_se_vazio(conn: sqlite3.Connection) -> bool:
    """Bootstrap idempotente da tabela bd_treinamentos a partir do xlsx empacotado.

    Popula se e somente se:
      - bd_treinamentos está vazia  E
      - registro_arquivos não tem entrada para tipo='treinamentos'

    Executa em transação única; rollback em falha. Retorna True se populou,
    False se já havia dados ou se o xlsx empacotado não estiver disponível.
    """
    logger.info('popular_treinamentos_se_vazio: verificando estado atual')
    row_count = conn.execute('SELECT COUNT(*) FROM bd_treinamentos').fetchone()[0]
    reg = conn.execute(
        "SELECT 1 FROM registro_arquivos WHERE tipo='treinamentos' LIMIT 1"
    ).fetchone()
    if row_count > 0 or reg is not None:
        logger.info('popular_treinamentos_se_vazio: já populado, no-op')
        return False

    xlsx = bundled_treinamentos_xlsx()
    if not xlsx.exists():
        logger.info('popular_treinamentos_se_vazio: xlsx empacotado ausente em %s', xlsx)
        return False

    logger.info('popular_treinamentos_se_vazio: bootstrap a partir de %s', xlsx)
    try:
        conn.execute('BEGIN')
        registrar_base_treinamentos(xlsx, conn)
    except Exception:
        conn.rollback()
        logger.exception('popular_treinamentos_se_vazio: rollback após falha')
        raise
    logger.info('popular_treinamentos_se_vazio: bootstrap concluído')
    return True


def obter_bd(conn: sqlite3.Connection) -> list[dict]:
    rows = conn.execute(
        'SELECT funcao, md_cobranca, area, quantidade FROM bd_distribuicao'
    ).fetchall()
    return [dict(r) for r in rows]


def obter_medicao(conn: sqlite3.Connection) -> list[dict]:
    rows = conn.execute(
        'SELECT data, sg_funcao, md_cobranca, pct_cobranca FROM medicao_frequencia'
    ).fetchall()
    return [dict(r) for r in rows]


def obter_registro_arquivos(conn: sqlite3.Connection) -> dict[str, dict]:
    rows = conn.execute(
        'SELECT tipo, caminho, importado_em FROM registro_arquivos'
    ).fetchall()
    return {r['tipo']: {'caminho': r['caminho'], 'importado_em': r['importado_em']} for r in rows}


def registrar_base_treinamentos(path: str | Path, conn: sqlite3.Connection) -> None:
    """Import Base de Treinamentos.xlsx into SQLite bd_treinamentos table."""
    logger.info('registrar_base_treinamentos: lendo %s', path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    records = []
    for row in wb.active.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        nome = str(row[0]).strip().upper()
        tipo_raw = str(row[1]).strip().lower() if row[1] else ''
        tipo = 'nao_remunerado' if ('não' in tipo_raw or 'nao' in tipo_raw) else 'remunerado'
        records.append((nome, tipo))
    wb.close()
    logger.info('registrar_base_treinamentos: %d treinamentos extraídos', len(records))

    conn.execute('DELETE FROM bd_treinamentos')
    conn.executemany(
        'INSERT INTO bd_treinamentos (nome, tipo) VALUES (?,?)',
        records,
    )
    conn.execute(
        'INSERT OR REPLACE INTO registro_arquivos (tipo, caminho, importado_em) VALUES (?,?,?)',
        ('treinamentos', str(path), datetime.now().isoformat()),
    )
    conn.commit()


def obter_tabela_treinamento(conn: sqlite3.Connection) -> dict[str, str]:
    """Return {nome_upper: tipo} from bd_treinamentos."""
    rows = conn.execute('SELECT nome, tipo FROM bd_treinamentos').fetchall()
    return {r['nome']: r['tipo'] for r in rows}
