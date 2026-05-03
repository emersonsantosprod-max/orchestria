"""
loaders.py — Leitura de fontes externas: treinamentos, férias e atestados.

Responsabilidade única: ler Excel e retornar dicts Python normalizados.
Não contém lógica de negócio nem conhecimento da ordem do pipeline.

Fonte = str | Path | BytesIO
  - str / Path → leitura de disco (testes, CLI legado)
  - BytesIO    → upload via FastAPI (sem tocar o filesystem)

openpyxl.load_workbook aceita ambos nativamente.
"""

from __future__ import annotations

import contextlib
import io
import logging
import os
import tempfile
import unicodedata
from pathlib import Path
from typing import Union

import openpyxl
import pandas as pd

from app.domain.errors import PlanilhaInvalidaError

logger = logging.getLogger(__name__)

# Tipo aceito por todas as funções públicas deste módulo
Fonte = Union[str, Path, io.BytesIO]

_COL_DATA_HR    = 0
_COL_RE_HR      = 1
_COL_HR_TRAB_HR = 19


def _is_path(fonte: Fonte) -> bool:
    return isinstance(fonte, (str, Path))


def _verificar_existe(fonte: Fonte) -> None:
    """Valida existência apenas para fontes de disco."""
    if _is_path(fonte) and not os.path.exists(fonte):
        raise FileNotFoundError(f"Arquivo não encontrado: {fonte}")


def carregar_dados_treinamento(fonte: Fonte) -> list[dict]:
    """Lê treinamentos realizados do xlsx mensal."""
    _verificar_existe(fonte)
    logger.info("carregar_dados_treinamento: carregando fonte")

    dados = []
    wb = openpyxl.load_workbook(fonte, read_only=True, data_only=True)
    for idx, row in enumerate(wb.active.iter_rows(min_row=3, values_only=True), start=3):
        if all(v is None for v in row):
            continue
        re_val, nome = row[0], row[1]
        if re_val is None and nome is None:
            continue
        dados.append({
            "linha":       idx,
            "matricula":   str(re_val).strip() if re_val is not None else "",
            "nome":        str(nome).strip()   if nome   is not None else "",
            "treinamento": str(row[5]).strip() if row[5] is not None else "",
            "data":        row[6],
            "carga":       str(row[7]).strip() if row[7] is not None else "",
        })
    wb.close()
    logger.info("carregar_dados_treinamento: %d registros lidos", len(dados))
    return dados


def carregar_dados_ferias_apenas(fonte: Fonte) -> list[dict]:
    """Lê só a planilha de férias (cabeçalho na linha 5)."""
    _verificar_existe(fonte)

    dados_ferias = []
    wb = openpyxl.load_workbook(fonte, read_only=True, data_only=True)
    for idx, row in enumerate(wb.active.iter_rows(min_row=6, values_only=True), start=6):
        if all(v is None for v in row):
            continue
        chapa = row[2] if len(row) > 2 else None
        if chapa is None:
            continue
        dados_ferias.append({
            "linha": idx,
            "chapa": chapa,
            "p1":    row[8]  if len(row) > 8  else None,
            "s1":    row[9]  if len(row) > 9  else None,
            "p2":    row[10] if len(row) > 10 else None,
            "s2":    row[11] if len(row) > 11 else None,
        })
    wb.close()
    return dados_ferias


def carregar_base_cobranca_xlsx(fonte: Fonte) -> dict[str, str]:
    """Lê base_cobranca → {sg_funcao_upper: md_cobranca_upper}."""
    _verificar_existe(fonte)

    base_cobranca: dict[str, str] = {}
    wb = openpyxl.load_workbook(fonte, read_only=True, data_only=True)
    for row in wb.active.iter_rows(values_only=True):
        if row[0] is None:
            continue
        chave = str(row[0]).strip().upper()
        valor = str(row[1]).strip().upper() if len(row) > 1 and row[1] is not None else ""
        if chave:
            base_cobranca[chave] = valor
    wb.close()
    return base_cobranca


def carregar_dados_ferias(
    fonte_ferias: Fonte,
    fonte_base_cobranca: Fonte,
) -> tuple[list[dict], dict[str, str]]:
    """Lê planilha de férias + tabela base_cobranca."""
    base_cobranca = carregar_base_cobranca_xlsx(fonte_base_cobranca)
    dados_ferias = carregar_dados_ferias_apenas(fonte_ferias)
    return dados_ferias, base_cobranca


def carregar_dados_atestado(fonte: Fonte) -> list:
    """Lê planilha de atestados médicos (Matrícula, Início, Fim).

    Conversão .xls → .xlsx aplica-se apenas a fontes de disco.
    BytesIO assumido xlsx — o caller é responsável pela conversão prévia.
    """
    _verificar_existe(fonte)

    # Conversão .xls apenas para paths de disco
    caminho_convertido: str | None = None
    if _is_path(fonte) and str(fonte).lower().endswith(".xls"):
        caminho_convertido = _converter_xls_para_xlsx(str(fonte))
        try:
            wb_tmp = openpyxl.load_workbook(caminho_convertido, read_only=True, data_only=True)
            wb_tmp.close()
        except Exception as err:
            raise ValueError("Falha ao converter arquivo .xls para .xlsx") from err
        fonte = caminho_convertido

    def _norm(s: str) -> str:
        s = unicodedata.normalize("NFKD", s)
        s = "".join(c for c in s if not unicodedata.combining(c))
        return " ".join(s.strip().lower().split())

    _EXATAS_MATRICULA = {"re", "matricula", "chapa"}
    _TOKENS_INICIO = [{"inicio"}, {"dt", "inicio"}, {"data", "inicio"}]
    _TOKENS_FIM = [{"fim"}, {"dt", "fim"}, {"data", "fim"}]

    def _match_inicio(h: str) -> bool:
        tokens = set(h.split())
        return any(t.issubset(tokens) for t in _TOKENS_INICIO)

    def _match_fim(h: str) -> bool:
        tokens = set(h.split())
        return any(t.issubset(tokens) for t in _TOKENS_FIM)

    try:
        wb = openpyxl.load_workbook(fonte, read_only=True, data_only=True)
        sheet = wb.active

        col_matricula = col_inicio = col_fim = None
        header_row = None

        for row_idx, row in enumerate(sheet.iter_rows(max_row=5, values_only=True), start=1):
            cols = {k: None for k in ("matricula", "inicio", "fim")}
            colisoes = {k: 0 for k in cols}

            for col_idx, cell in enumerate(row):
                if cell is None or not isinstance(cell, str):
                    continue
                h = _norm(cell)
                if h in _EXATAS_MATRICULA:
                    cols["matricula"] = col_idx
                    colisoes["matricula"] += 1
                if _match_inicio(h):
                    cols["inicio"] = col_idx
                    colisoes["inicio"] += 1
                if _match_fim(h):
                    cols["fim"] = col_idx
                    colisoes["fim"] += 1

            if any(v > 1 for v in colisoes.values()):
                continue
            if all(v is not None for v in cols.values()):
                col_matricula = cols["matricula"]
                col_inicio = cols["inicio"]
                col_fim = cols["fim"]
                header_row = row_idx
                break

        if col_matricula is None or col_inicio is None or col_fim is None:
            wb.close()
            raise ValueError(
                "Arquivo de atestados não contém as colunas obrigatórias: "
                "Matrícula, Início, Fim."
            )

        dados = []
        for idx, row in enumerate(
            sheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            if all(v is None for v in row):
                continue
            mat = row[col_matricula]
            ini = row[col_inicio]
            fim = row[col_fim]
            if mat is None or ini is None or fim is None:
                continue
            dados.append({"linha": idx, "matricula": mat, "inicio": ini, "fim": fim})

        wb.close()
        return dados
    finally:
        if caminho_convertido:
            with contextlib.suppress(Exception):
                os.remove(caminho_convertido)


def carregar_medicao_hr(fonte: Fonte) -> tuple[list[dict], int]:
    """Lê a sheet 'Frequencia' e extrai (matricula, data, hr_trabalhadas).

    Retorna (registros, n_linhas_dados).
    """
    _verificar_existe(fonte)

    wb = openpyxl.load_workbook(fonte, read_only=True, data_only=True)
    try:
        if "Frequencia" in wb.sheetnames:
            ws = wb["Frequencia"]
        elif "Frequência" in wb.sheetnames:
            ws = wb["Frequência"]
        else:
            raise PlanilhaInvalidaError(
                "Planilha de Medição não contém a aba 'Frequencia'."
            )

        registros: list[dict] = []
        n = 0
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            if row is None or all(c is None for c in row):
                continue
            n += 1
            mat_val = row[_COL_RE_HR]      if len(row) > _COL_RE_HR      else None
            dat_val = row[_COL_DATA_HR]    if len(row) > _COL_DATA_HR    else None
            hr_val  = row[_COL_HR_TRAB_HR] if len(row) > _COL_HR_TRAB_HR else None
            matricula = str(mat_val).strip() if mat_val is not None else ""
            if hasattr(dat_val, "strftime"):
                data_str = dat_val.strftime("%d/%m/%Y")
            else:
                data_str = str(dat_val).strip() if dat_val is not None else ""
            hr_trab = float(hr_val) if hr_val is not None else None
            registros.append({
                "matricula": matricula,
                "data": data_str,
                "hr_trabalhadas": hr_trab,
            })
    finally:
        wb.close()
    return registros, n


def _converter_xls_para_xlsx(caminho_xls: str) -> str:
    """Converte .xls para .xlsx temporário. Retorna caminho do arquivo gerado."""
    dfs = pd.read_excel(caminho_xls, sheet_name=None, engine="xlrd", header=None)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
    with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer_:
        for sheet_name, df in dfs.items():
            df.to_excel(writer_, sheet_name=sheet_name, index=False, header=False)
    return tmp_path
