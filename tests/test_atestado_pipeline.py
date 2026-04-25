import os
from datetime import date

import openpyxl

from app import db, pipeline


def _criar_atestado_xlsx(caminho: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Matricula', 'Inicio', 'Fim'])
    ws.append(['111', date(2026, 3, 18), date(2026, 3, 19)])
    wb.save(caminho)
    wb.close()


def test_pipeline_atestado_aplica_observacao(tmp_path):
    base_dir = os.path.dirname(os.path.abspath(__file__))
    medicao = os.path.join(base_dir, 'fixtures', 'medicao_mock.xlsx')
    saida = str(tmp_path / 'saida_atestado.xlsx')
    atestado_path = str(tmp_path / 'atestado_mock.xlsx')
    _criar_atestado_xlsx(atestado_path)

    conn = db.conectar(str(tmp_path / 'test.db'))
    try:
        resultado = pipeline.executar_pipeline(
            caminho_medicao=medicao,
            caminho_atestado=atestado_path,
            caminho_saida=saida,
            conn=conn,
        )
    finally:
        conn.close()

    assert resultado.atestados_processados == 1
    assert resultado.atestados_atualizados == 2

    wb = openpyxl.load_workbook(saida)
    ws = wb['Frequência'] if 'Frequência' in wb.sheetnames else wb['Frequencia']
    assert ws.cell(row=2, column=3).value == 'ATESTADO MÉDICO'
    assert ws.cell(row=3, column=3).value == 'ATESTADO MÉDICO'
    wb.close()
