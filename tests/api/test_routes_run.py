"""Testes de POST /api/run/{ferias,atestado,distribuicao}."""

from __future__ import annotations

import io
import sqlite3
from datetime import date
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.infrastructure.data import (
    DistribuicaoRepository,
    FeriasRepository,
    conectar,
)

FIXTURES = Path(__file__).resolve().parents[1] / "fixtures"
MEDICAO_MOCK = FIXTURES / "medicao_mock.xlsx"


@pytest.fixture
def client(tmp_path):
    from app.api.dependencies import get_conn
    from app.api.main import app

    db = tmp_path / "test.db"
    conectar(db).close()

    def _override():
        c = sqlite3.connect(str(db), timeout=5, check_same_thread=False)
        c.row_factory = sqlite3.Row
        try:
            yield c
        finally:
            c.close()

    app.dependency_overrides[get_conn] = _override
    try:
        yield TestClient(app), db
    finally:
        app.dependency_overrides.clear()
        import logging
        for h in logging.root.handlers[:]:
            if getattr(h, "_automacao_file_handler", False):
                logging.root.removeHandler(h)
                h.close()


def _seed_cobranca(db_path: Path) -> None:
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    try:
        FeriasRepository(conn).salvar([("ENC", "MEC", 1)])
        conn.commit()
    finally:
        conn.close()


def _seed_distribuicao(db_path: Path) -> None:
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    try:
        DistribuicaoRepository(conn).salvar([("ENC", "MEC", "AREA1", 1)])
        conn.commit()
    finally:
        conn.close()


def _atestado_xlsx_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Matricula", "Inicio", "Fim"])
    ws.append(["111", date(2026, 3, 18), date(2026, 3, 19)])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _ferias_xlsx_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["chapa", "p1", "s1", "p2", "s2"])
    ws.append(["1.000111", "18/03/2026 a 19/03/2026", "Aprovado", None, None])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _medicao_para_ferias_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Frequência"
    ws.append(["Data", "RE", "Observação", "Descontos", "Situação", "Md Cobrança", "Sg Função"])
    ws.append(["18/03/2026", "111", None, None, None, "MEC", "ENC"])
    ws.append(["19/03/2026", "111", None, None, None, "MEC", "ENC"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_run_atestado_processa_relatorio(client):
    tc, _ = client
    res = tc.post(
        "/api/run/atestado",
        files={
            "medicao": ("medicao.xlsx", MEDICAO_MOCK.read_bytes()),
            "relatorio": ("atestado.xlsx", _atestado_xlsx_bytes()),
        },
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["processados"] == 1
    assert body["atualizados"] == 2
    assert body["arquivo_saida"] == "medicao_atestado_processada.xlsx"


def test_run_ferias_sem_cobranca_retorna_422(client):
    tc, _ = client
    res = tc.post(
        "/api/run/ferias",
        files={
            "medicao": ("medicao.xlsx", MEDICAO_MOCK.read_bytes()),
            "relatorio": ("ferias.xlsx", _ferias_xlsx_bytes()),
        },
    )
    assert res.status_code == 422
    assert "bd_cobranca" in res.json()["detail"]


def test_run_ferias_com_cobranca_executa_pipeline(client):
    tc, db = client
    _seed_cobranca(db)
    res = tc.post(
        "/api/run/ferias",
        files={
            "medicao": ("medicao.xlsx", _medicao_para_ferias_bytes()),
            "relatorio": ("ferias.xlsx", _ferias_xlsx_bytes()),
        },
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["arquivo_saida"] == "medicao_ferias_processada.xlsx"
    assert isinstance(body["inconsistencias"], list)


def test_run_distribuicao_sem_bd_retorna_422(client):
    tc, _ = client
    res = tc.post(
        "/api/run/distribuicao",
        files={"medicao": ("medicao.xlsx", MEDICAO_MOCK.read_bytes())},
    )
    assert res.status_code == 422
    assert "bd_distribuicao" in res.json()["detail"]


def test_run_distribuicao_com_bd_executa_validacao(client):
    tc, db = client
    _seed_distribuicao(db)
    res = tc.post(
        "/api/run/distribuicao",
        files={"medicao": ("medicao.xlsx", MEDICAO_MOCK.read_bytes())},
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["processados"] == 0
    assert body["atualizados"] == 0
    assert body["arquivo_saida"] == "medicao_distribuicao_processada.xlsx"
