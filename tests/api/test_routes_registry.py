"""Testes das rotas /api/registry/<tipo> (path-based, 4a)."""

from __future__ import annotations

import shutil
import sqlite3
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.infrastructure.data import conectar

FIXTURES = Path(__file__).resolve().parents[1] / "fixtures"
MEDICAO_MOCK = FIXTURES / "medicao_mock.xlsx"


@pytest.fixture
def client(isolated_paths):
    from app.api.dependencies import get_conn
    from app.api.main import app

    db = isolated_paths / "test.db"
    conectar(db).close()

    def _override():
        c = sqlite3.connect(str(db), timeout=5, check_same_thread=False)
        c.row_factory = sqlite3.Row
        try:
            yield c
        finally:
            c.close()

    app.dependency_overrides[get_conn] = _override
    try:
        yield TestClient(app), isolated_paths
    finally:
        app.dependency_overrides.clear()


def _xlsx_treinamentos(tmp: Path) -> Path:
    p = tmp / "treinos.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Nome", "Tipo"])
    ws.append(["NR-10", "Remunerado"])
    wb.save(p)
    return p


def _xlsx_cobranca(tmp: Path) -> Path:
    p = tmp / "cobranca.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ENC", "FÉRIAS NORMAIS"])
    wb.save(p)
    return p


def test_registry_medicao_happy(client):
    c, tmp = client
    # Copiar fixture para o tmp_path
    p = tmp / "medicao.xlsx"
    shutil.copy(MEDICAO_MOCK, p)
    r = c.post("/api/registry/medicao", json={"caminho": str(p)})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["caminho"] == str(p)
    assert len(body["mes_referencia"]) == 7


def test_registry_medicao_caminho_inexistente_404(client):
    c, tmp = client
    r = c.post("/api/registry/medicao", json={"caminho": str(tmp / "no.xlsx")})
    assert r.status_code == 404
    assert "ARQUIVO_NAO_ENCONTRADO" in r.json()["detail"]


def test_registry_medicao_extensao_invalida_422(client):
    c, tmp = client
    p = tmp / "arquivo.txt"
    p.write_text("x")
    r = c.post("/api/registry/medicao", json={"caminho": str(p)})
    assert r.status_code == 422


def test_registry_treinamentos_happy(client):
    c, tmp = client
    p = _xlsx_treinamentos(tmp)
    r = c.post("/api/registry/treinamentos", json={"caminho": str(p)})
    assert r.status_code == 200, r.text
    assert r.json()["qtd"] >= 1


def test_registry_cobranca_happy(client):
    c, tmp = client
    p = _xlsx_cobranca(tmp)
    r = c.post("/api/registry/cobranca", json={"caminho": str(p)})
    assert r.status_code == 200, r.text
    assert r.json()["qtd"] >= 1


def test_registry_cobranca_arquivo_vazio_422(client):
    c, tmp = client
    p = tmp / "vazio.xlsx"
    wb = openpyxl.Workbook()  # sem linhas — registrar_cobranca exige ≥1 row util
    wb.save(p)
    r = c.post("/api/registry/cobranca", json={"caminho": str(p)})
    assert r.status_code == 422


def test_registry_get_consulta(client):
    c, tmp = client
    # Sem nada registrado
    r = c.get("/api/registry/medicao")
    assert r.status_code == 200
    assert r.json()["registrado"] is False

    # Após registrar
    p = tmp / "medicao.xlsx"
    shutil.copy(MEDICAO_MOCK, p)
    c.post("/api/registry/medicao", json={"caminho": str(p)})
    r = c.get("/api/registry/medicao")
    body = r.json()
    assert body["registrado"] is True
    assert body["caminho"] == str(p)
