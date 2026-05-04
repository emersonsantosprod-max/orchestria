"""Testes de POST /api/config/catalogo e POST /api/config/medicao."""

from __future__ import annotations

import io
import sqlite3
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.infrastructure.data import conectar

FIXTURES = Path(__file__).resolve().parents[1] / "fixtures"


@pytest.fixture
def client(tmp_path):
    from app.api.dependencies import get_conn
    from app.api.main import app

    db = tmp_path / "test.db"
    conectar(db).close()  # cria schema

    def _override():
        c = sqlite3.connect(str(db), timeout=5, check_same_thread=False)
        c.row_factory = sqlite3.Row
        try:
            yield c
        finally:
            c.close()

    app.dependency_overrides[get_conn] = _override
    try:
        yield TestClient(app)
    finally:
        app.dependency_overrides.clear()
        import logging
        for h in logging.root.handlers[:]:
            if getattr(h, "_automacao_file_handler", False):
                logging.root.removeHandler(h)
                h.close()


def _xlsx_bytes(path: Path) -> bytes:
    return path.read_bytes()


def _xlsx_invalido() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Frequencia"
    ws.append(["coluna_estranha", "outra"])
    ws.append(["x", "y"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_medicao_valido() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Frequencia"
    ws.append(["data", "sg funcao", "md cobranca", "% cobrança"])
    ws.append(["01/01/2026", "ENC", "MEC", 1.0])
    ws.append(["02/01/2026", "ENC", "MEC", 0.5])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_medicao_multi_mes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Frequencia"
    ws.append(["data", "sg funcao", "md cobranca", "% cobrança"])
    ws.append(["01/01/2026", "ENC", "MEC", 1.0])
    ws.append(["01/02/2026", "ENC", "MEC", 1.0])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_upload_catalogo_popula_e_initial_data_fica_ready(client):
    res = client.post(
        "/api/config/catalogo",
        files={"arquivo": ("treinamentos.xlsx", _xlsx_bytes(FIXTURES / "base_treinamentos_mock.xlsx"))},
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["count"] > 0
    assert body["arquivo"] == "treinamentos.xlsx"

    initial = client.get("/api/initial-data").json()
    assert initial["catalog_status"] == "CATALOG_READY"


def test_upload_medicao_popula_e_initial_data_fica_ready(client):
    res = client.post(
        "/api/config/medicao",
        files={"arquivo": ("medicao.xlsx", _xlsx_medicao_valido())},
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["count"] > 0
    assert body["arquivo"] == "medicao.xlsx"
    assert isinstance(body["avisos"], list)

    initial = client.get("/api/initial-data").json()
    assert initial["measurement_status"] == "MEASUREMENT_READY"
    assert initial["mes_referencia"] == "2026-01"


def test_initial_data_sem_medicao_tem_mes_referencia_none(client):
    initial = client.get("/api/initial-data").json()
    assert initial["measurement_status"] == "MEASUREMENT_MISSING"
    assert initial["mes_referencia"] is None


def test_initial_data_multi_mes_retorna_mes_referencia_none(client):
    res = client.post(
        "/api/config/medicao",
        files={"arquivo": ("medicao.xlsx", _xlsx_medicao_multi_mes())},
    )
    assert res.status_code == 200, res.text

    initial = client.get("/api/initial-data").json()
    assert initial["measurement_status"] == "MEASUREMENT_READY"
    assert initial["mes_referencia"] is None


def test_upload_medicao_xlsx_invalido_retorna_422(client):
    res = client.post(
        "/api/config/medicao",
        files={"arquivo": ("ruim.xlsx", _xlsx_invalido())},
    )
    assert res.status_code == 422
    assert "Cabeçalho" in res.json()["detail"] or "cabeçalho" in res.json()["detail"].lower()


def test_upload_sem_arquivo_retorna_422(client):
    res = client.post("/api/config/catalogo")
    assert res.status_code == 422
