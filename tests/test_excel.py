import re
import zipfile

import openpyxl
import pytest

from app import excel as writer
from app.core import Update


def criar_planilha_mock(caminho):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Frequência'
    ws.append(['Data', 'RE', 'Observação', 'Descontos'])
    ws.append(['18/03/2026', '123', '', ''])
    wb.save(caminho)


def test_localizar_linhas_e_colunas(tmp_path):
    caminho = str(tmp_path / "teste_writer.xlsx")
    criar_planilha_mock(caminho)

    _, sheet = writer.carregar_planilha(caminho)
    col_map = writer.mapear_colunas(sheet)

    assert col_map['data'] == 0
    assert col_map['matricula'] == 1
    assert col_map['observacao'] == 2
    assert col_map['desconto'] == 3

    index = writer.indexar_e_ler_dados(sheet, col_map)[0]
    assert ('123', '18/03/2026') in index
    assert index[('123', '18/03/2026')] == [2]


def test_erro_matricula_nao_encontrada(tmp_path):
    caminho = str(tmp_path / "teste_writer_erro.xlsx")
    criar_planilha_mock(caminho)

    _, sheet = writer.carregar_planilha(caminho)
    col_map = writer.mapear_colunas(sheet)
    index = writer.indexar_e_ler_dados(sheet, col_map)[0]

    updates = [Update(
        tipo='treinamento',
        matricula='404',
        data='18/03/2026',
        observacao='T',
        desconto_min=0,
        sobrescrever_obs=True,
    )]

    _, erros = writer.aplicar_updates(updates, col_map, index)

    assert len(erros) == 1
    assert erros[0].matricula == '404'
    assert erros[0].erro == 'matrícula não encontrada'
    assert erros[0].origem == 'writer'


def test_mapear_ambiguidade_descontos(tmp_path):
    """Dois cabeçalhos que casariam com 'desconto' devem ser rejeitados."""
    caminho = str(tmp_path / "ambiguo.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Frequência'
    ws.append(['Data', 'RE', 'Descontos', 'Descontos Extras', 'Observação'])
    wb.save(caminho)

    _, sheet = writer.carregar_planilha(caminho)
    with pytest.raises(ValueError, match=r"\[MAPEAMENTO\] Ambiguidade"):
        writer.mapear_colunas(sheet)


def test_mapear_com_coluna_nova_no_meio(tmp_path):
    """Coluna nova inserida no meio (como Equipe em April) não quebra mapeamento."""
    caminho = str(tmp_path / "abril.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Frequência'
    ws.append(['Data', 'RE', 'Supervisor', 'Encarregado', 'Equipe', 'Ronda',
               'Apoio Calculo Descanso', 'Desconto Descanso', 'Descontos',
               'Observação', 'Total Descontos'])
    wb.save(caminho)

    _, sheet = writer.carregar_planilha(caminho)
    col_map = writer.mapear_colunas(sheet)

    assert col_map['data'] == 0
    assert col_map['matricula'] == 1
    assert col_map['desconto'] == 8
    assert col_map['observacao'] == 9
    assert col_map['_header_row'] == 1


def test_aplicar_updates_gera_patches(tmp_path):
    caminho = str(tmp_path / "teste_writer_sucesso.xlsx")
    criar_planilha_mock(caminho)

    _, sheet = writer.carregar_planilha(caminho)
    col_map = writer.mapear_colunas(sheet)
    index, _, descontos_existentes, *_ = writer.indexar_e_ler_dados(sheet, col_map)

    updates = [Update(
        tipo='treinamento',
        matricula='123',
        data='18/03/2026',
        observacao='NOVA OBS',
        desconto_min=120,
        sobrescrever_obs=True,
    )]

    patches, erros = writer.aplicar_updates(
        updates, col_map, index,
        descontos_existentes=descontos_existentes,
    )

    assert erros == []
    col_obs_1 = col_map['observacao'] + 1
    col_desc_1 = col_map['desconto'] + 1
    assert patches[(2, col_obs_1)] == 'NOVA OBS'
    assert patches[(2, col_desc_1)] == '02:00'


def _criar_xlsx_com_namespaces_mc(caminho: str) -> None:
    """Cria um xlsx mínimo cujo sheet2.xml declara xmlns:x14ac e mc:Ignorable."""
    # Gera base via openpyxl para estrutura válida de zip
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Frequência'
    ws.append(['Data', 'RE', 'Observação', 'Descontos'])
    ws.append(['18/03/2026', '123', '', ''])
    wb.save(caminho)

    # Substitui xl/worksheets/sheet1.xml por XML com namespaces MC completos
    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<worksheet'
        ' xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
        ' xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006"'
        ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'
        ' xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"'
        ' xmlns:xr="http://schemas.microsoft.com/office/spreadsheetml/2014/revision"'
        ' xmlns:xr2="http://schemas.microsoft.com/office/spreadsheetml/2015/revision2"'
        ' xmlns:xr3="http://schemas.microsoft.com/office/spreadsheetml/2016/revision3"'
        ' mc:Ignorable="x14ac xr xr2 xr3">'
        '<sheetData>'
        '<row r="1"><c r="A1" t="inlineStr"><is><t>Data</t></is></c>'
        '<c r="B1" t="inlineStr"><is><t>RE</t></is></c>'
        '<c r="C1" t="inlineStr"><is><t>Observação</t></is></c>'
        '<c r="D1" t="inlineStr"><is><t>Descontos</t></is></c></row>'
        '<row r="2"><c r="A2" t="inlineStr"><is><t>18/03/2026</t></is></c>'
        '<c r="B2" t="inlineStr"><is><t>123</t></is></c>'
        '<c r="C2" t="inlineStr"><is><t></t></is></c>'
        '<c r="D2" t="inlineStr"><is><t></t></is></c></row>'
        '</sheetData></worksheet>'
    ).encode()

    # Reescreve o zip injetando o sheet XML com namespaces completos
    import io
    buf = io.BytesIO()
    with zipfile.ZipFile(caminho) as src_z, zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as dst_z:
        for info in src_z.infolist():
            if info.filename.endswith('sheet1.xml'):
                dst_z.writestr(info, sheet_xml)
            else:
                dst_z.writestr(info, src_z.read(info.filename))
    with open(caminho, 'wb') as f:
        f.write(buf.getvalue())


def test_salvar_via_zip_preserva_namespaces_mc(tmp_path):
    """salvar_via_zip não deve perder xmlns:x14ac nem qualquer token de mc:Ignorable."""
    src = str(tmp_path / 'src.xlsx')
    dst = str(tmp_path / 'dst.xlsx')
    _criar_xlsx_com_namespaces_mc(src)

    # patches vazios — só queremos verificar que namespaces são preservados
    writer.salvar_via_zip(src, dst, {}, nome_aba='Frequência')

    with zipfile.ZipFile(dst) as z:
        # Encontra o sheet da aba Frequência
        sheet_names = [n for n in z.namelist() if re.match(r'xl/worksheets/sheet\d+\.xml', n)]
        assert sheet_names, 'nenhum sheet encontrado'
        xml_str = z.read(sheet_names[0]).decode('utf-8')

    # Todos os xmlns devem estar presentes
    for prefix in ('x14ac', 'xr', 'xr2', 'xr3', 'mc'):
        assert f'xmlns:{prefix}=' in xml_str, f'xmlns:{prefix} ausente no sheet reescrito'

    # mc:Ignorable deve estar intacto
    assert 'mc:Ignorable="x14ac xr xr2 xr3"' in xml_str

    # Invariante genérica: todo token em mc:Ignorable tem xmlns: correspondente
    ignorable_match = re.search(r'mc:Ignorable="([^"]+)"', xml_str)
    assert ignorable_match, 'mc:Ignorable não encontrado'
    for token in ignorable_match.group(1).split():
        assert f'xmlns:{token}=' in xml_str, f'token {token!r} em mc:Ignorable sem xmlns declarado'
