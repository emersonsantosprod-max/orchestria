"""
Equivalência legado ↔ nova pipeline (adapter + domínio).

Disposable: removido no Step 6, junto com o módulo legado. Garante que
``app.distribuicao_contratual`` e
``app.infrastructure.adapters.excel_distribuicao_contratual`` +
``app.domain.distribuicao_contratual`` produzem o mesmo resultado para a
mesma entrada — comparação **semântica** do xlsx (linhas como tuplas), não
byte-a-byte.
"""

from __future__ import annotations

import openpyxl
import pytest

from app import distribuicao_contratual as legacy
from app.domain.distribuicao_contratual import (
    localizar_colunas_chave,
    normalizar_linhas,
    parse_distribuicao_cols,
    validar_distribuicao_cobranca,
)
from app.infrastructure.adapters.excel_distribuicao_contratual import (
    escrever_xlsx_normalizado,
    ler_xlsx_contratual,
)

_FIXTURE_ROWS = [
    [
        'TP MO', 'ÁREA', 'FUNÇÃO', 'SIGLA',
        'CENTRAL', 'ADM-B',
        'BREAKDOWN PE1', 'BREAKDOWN PVC',
        'HD PVC', 'CV UA', 'ANALITICA',
        'COLUNA_ESTRANHA',
        'Atual', 'OBSERVAÇÕES',
    ],
    ['MOD', 'ELÉTRICA', 'ELETRICISTA I',  'ELET-I',  5, 1, 2, 1, None, None, None, None, 9, ''],
    ['MOD', 'MECÂNICA', 'MECANICO II',    'MEC-II',  3, 0, 0, 0, 1.5,  0,    None, None, 5, ''],
    ['MOD', 'INSTRUM',  'INSTRUMENTISTA', 'INST-I',  0, 0, 0, 0, 0,    2,    1,    None, 4, ''],
    ['MOD', 'GERAL',    'AJUDANTE',       None,      0, 0, 0, 0, 0,    0,    1,    None, 0, ''],
]


@pytest.fixture(scope='module')
def fixture_path(tmp_path_factory) -> str:
    path = tmp_path_factory.mktemp('eq_dc') / 'entrada.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in _FIXTURE_ROWS:
        ws.append(row)
    wb.save(path)
    wb.close()
    return str(path)


def _via_dominio(path):
    headers, data_rows, w_hdr = ler_xlsx_contratual(path)
    col_map, w0 = parse_distribuicao_cols(headers)
    sigla_col, funcao_col, atual_col = localizar_colunas_chave(headers)
    normalized, raw_sums, atual, w1 = normalizar_linhas(
        data_rows, col_map, sigla_col, funcao_col, atual_col)
    return normalized, raw_sums, atual, w_hdr + w0 + w1


def _key_norm(r): return (r['funcao'], r['md_cobranca'], r['area'] or '')
def _key_warn(w): return (w.get('tipo', ''), w.get('funcao', ''), w.get('coluna', ''),
                          w.get('md_cobranca', ''), w.get('area') or '')


def test_normalizar_equivalente(fixture_path):
    n_l, raw_l, atual_l, w_l = legacy.carregar_e_normalizar(fixture_path)
    n_d, raw_d, atual_d, w_d = _via_dominio(fixture_path)

    assert sorted(n_l, key=_key_norm) == sorted(n_d, key=_key_norm)
    assert raw_l == raw_d
    assert atual_l == atual_d
    assert sorted(w_l, key=_key_warn) == sorted(w_d, key=_key_warn)


def test_validar_equivalente(fixture_path):
    n_l, raw_l, atual_l, _ = legacy.carregar_e_normalizar(fixture_path)
    n_d, raw_d, atual_d, _ = _via_dominio(fixture_path)
    inc_l = legacy.validar_distribuicao_cobranca(n_l, raw_l, atual_l)
    inc_d = validar_distribuicao_cobranca(n_d, raw_d, atual_d)
    assert sorted(inc_l, key=_key_warn) == sorted(inc_d, key=_key_warn)


def _read_rows(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = [tuple(r) for r in wb.active.iter_rows(values_only=True)]
    wb.close()
    return rows


def test_xlsx_normalizado_equivalente_semanticamente(fixture_path, tmp_path):
    n_l, _, _, _ = legacy.carregar_e_normalizar(fixture_path)
    n_d, _, _, _ = _via_dominio(fixture_path)

    legacy_out = tmp_path / 'legacy.xlsx'
    new_out = tmp_path / 'new.xlsx'
    legacy.exportar_normalizado(n_l, str(legacy_out))
    escrever_xlsx_normalizado(n_d, new_out)

    assert _read_rows(legacy_out) == _read_rows(new_out)
