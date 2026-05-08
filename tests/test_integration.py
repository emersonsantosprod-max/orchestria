import os

import openpyxl

from app.application import pipeline as service
from app.infrastructure import data


def test_fluxo_completo(tmp_path):
    """
    Testa todo o ciclo usando apenas as fixtures geradas e valida
    a saída e a planilha final sem dependências externas reais.
    """
    base_dir = os.path.dirname(os.path.abspath(__file__))
    f_dir = os.path.join(base_dir, 'fixtures')

    medicao = os.path.join(f_dir, 'medicao_mock.xlsx')
    trein = os.path.join(f_dir, 'treinamentos_mock.xlsx')
    base_tr = os.path.join(f_dir, 'base_treinamentos_mock.xlsx')
    saida = str(tmp_path / "saida_mock.xlsx")
    db_file = str(tmp_path / "test.db")

    conn = data.conectar(db_file)
    data.popular_bd_se_vazio(conn)
    data.registrar_base_treinamentos(base_tr, conn)
    try:
        res = service.executar_pipeline(
            caminho_medicao=medicao,
            caminho_treinamentos=trein,
            caminho_ferias='',
            caminho_base_cobranca='',
            caminho_atestado='',
            caminho_saida=saida,
            conn=conn,
            validar_distribuicao=True,
        )
    finally:
        conn.close()

    assert res.processados == 3
    assert res.atualizados == 3
    assert len(res.inconsistencias) == 1
    assert 'validar_distribuicao=True ignorado' in res.inconsistencias[0].erro

    wb = openpyxl.load_workbook(saida)
    ws = wb['Frequência']

    assert ws.cell(row=2, column=3).value == "TREIN. TR-SIMPLES - 2H; TREIN. TR-REMUNERADO - 8H (NÃO DESCONTA)"
    assert ws.cell(row=2, column=4).value == "01:10"

    assert ws.cell(row=4, column=3).value == "TREIN. TR-MULTIDIA - 8H"
    assert ws.cell(row=4, column=4).value == "08:00"

    assert ws.cell(row=5, column=3).value == "TREIN. TR-MULTIDIA - 8H"
    assert ws.cell(row=5, column=4).value == "08:00"

    wb.close()
