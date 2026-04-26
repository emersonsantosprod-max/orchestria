"""Pipeline state-based test: férias sobrescreve treinamento na mesma célula.

Constrói fixtures xlsx mínimas em tmp_path, executa
pipeline.executar_pipeline com treinamento + férias ativos para uma
mesma (matrícula, data), abre o xlsx de saída e assere que o valor
final da célula de observação é a observação de férias.

Pinniza o contrato observável de CLAUDE.md → CRITICAL ("ordem
treinamento + férias garante que sobrescrever_obs=True de férias
vença"). State-based: independente de como o pipeline compõe
internamente a lista de updates — sobrevive a refactors futuros.

Permanente — não deletar.
"""

from __future__ import annotations

import os

import openpyxl

from app.application import pipeline
from app.infrastructure import db


def _criar_medicao_xlsx(caminho: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Frequência'
    ws.append(['Data', 'RE', 'Observação', 'Descontos',
               'Situação', 'MD Cobrança', 'Sg Função'])
    ws.append(['18/03/2026', '111', '', '', '', 'ADICIONAL', 'X'])
    wb.save(caminho)
    wb.close()


def _criar_treinamento_xlsx(caminho: str, nome_treinamento: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['RE', 'Nome', 'a', 'b', 'c', 'Treinamento', 'Data', 'Carga'])
    ws.append(['ignored', 'header', None, None, None, None, None, None])
    ws.append(['111', 'FULANO', None, None, None, nome_treinamento,
               '18/03/2026', '2H'])
    wb.save(caminho)
    wb.close()


def _criar_base_treinamentos_xlsx(caminho: str, nome_treinamento: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['TREINAMENTO', 'TIPO', 'CARGA'])
    ws.append([nome_treinamento, 'Remunerado', '2H'])
    wb.save(caminho)
    wb.close()


def _criar_ferias_xlsx(caminho: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    for _ in range(5):
        ws.append([None] * 12)
    ws.append([
        None, None, '1.000111', None, None, None, None, None,
        '18/03/2026 a 18/03/2026', 'Aprovado', None, None,
    ])
    wb.save(caminho)
    wb.close()


def _criar_base_cobranca_xlsx(caminho: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['X', 'FÉRIAS'])
    wb.save(caminho)
    wb.close()


def test_ferias_overwrites_treinamento_on_same_cell_via_pipeline(tmp_path):
    medicao = str(tmp_path / 'medicao.xlsx')
    treinamento = str(tmp_path / 'treinamento.xlsx')
    base_treinamentos = str(tmp_path / 'base_treinamentos.xlsx')
    ferias = str(tmp_path / 'ferias.xlsx')
    base_cobranca = str(tmp_path / 'base_cobranca.xlsx')
    saida = str(tmp_path / 'saida.xlsx')
    db_file = str(tmp_path / 'test.db')

    nome_treinamento = 'TR-NORMAL'
    _criar_medicao_xlsx(medicao)
    _criar_treinamento_xlsx(treinamento, nome_treinamento)
    _criar_base_treinamentos_xlsx(base_treinamentos, nome_treinamento)
    _criar_ferias_xlsx(ferias)
    _criar_base_cobranca_xlsx(base_cobranca)

    conn_seed = db.conectar(db_file)
    db.registrar_base_treinamentos(base_treinamentos, conn_seed)
    conn_seed.close()

    conn = db.conectar(db_file)
    try:
        resultado = pipeline.executar_pipeline(
            caminho_medicao=medicao,
            caminho_treinamentos=treinamento,
            caminho_ferias=ferias,
            caminho_base_cobranca=base_cobranca,
            caminho_saida=saida,
            conn=conn,
        )
    finally:
        conn.close()

    assert os.path.exists(saida)
    assert resultado.processados == 1
    assert resultado.ferias_processadas == 1

    wb = openpyxl.load_workbook(saida)
    ws = wb['Frequência']
    cell_observacao = ws.cell(row=2, column=3).value
    cell_situacao = ws.cell(row=2, column=5).value
    wb.close()

    assert cell_observacao == '18/03 a 18/03 - FÉRIAS', (
        f"férias deveria sobrescrever treinamento; obs final = {cell_observacao!r}"
    )
    assert cell_situacao == 'FÉRIAS'
