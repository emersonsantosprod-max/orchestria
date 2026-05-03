"""Adapter tests: round-trip, determinismo, erros estruturais."""

from __future__ import annotations

import openpyxl
import pytest

from app.infrastructure.excel_distribuicao import (
    DistribuicaoContratualMalformadaError,
    escrever_xlsx_normalizado,
    ler_xlsx_contratual,
)


def _make_xlsx(rows: list, path) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    wb.save(path)
    wb.close()
    return str(path)


def test_round_trip_escrever_e_ler(tmp_path):
    records = [
        {'funcao': 'ELET-I', 'md_cobranca': 'CENTRAL',   'area': None,  'quantidade': 5},
        {'funcao': 'ELET-I', 'md_cobranca': 'BREAKDOWN', 'area': 'PVC', 'quantidade': 2},
    ]
    out_path = tmp_path / 'norm.xlsx'
    escrever_xlsx_normalizado(records, out_path)

    wb = openpyxl.load_workbook(out_path, read_only=True, data_only=True)
    rows = [tuple(r) for r in wb.active.iter_rows(values_only=True)]
    wb.close()

    assert rows[0] == ('funcao', 'md_cobranca', 'area', 'quantidade')
    assert {tuple(r) for r in rows[1:]} == {
        ('ELET-I', 'CENTRAL', None, 5),
        ('ELET-I', 'BREAKDOWN', 'PVC', 2),
    }


def test_ler_xlsx_contratual_basico(tmp_path):
    path = _make_xlsx(
        [
            ['preâmbulo', None, None],
            ['SIGLA', 'CENTRAL', 'Atual'],
            ['ELET-I', 5, 5],
        ],
        tmp_path / 'in.xlsx',
    )
    headers, data_rows, header_warnings = ler_xlsx_contratual(path)
    assert headers == ('SIGLA', 'CENTRAL', 'Atual')
    assert data_rows == [('ELET-I', 5, 5)]
    assert header_warnings == []


def test_ler_xlsx_contratual_emite_aviso_header_duplicado(tmp_path):
    path = _make_xlsx(
        [
            ['SIGLA', 'CENTRAL'],
            ['ELET-I', 5],
            ['SIGLA', 'CENTRAL'],
            ['ELET-II', 3],
        ],
        tmp_path / 'in.xlsx',
    )
    _, _, header_warnings = ler_xlsx_contratual(path)
    assert any(w['tipo'] == 'AVISO_HEADER_DUPLICADO' for w in header_warnings)


def test_ler_xlsx_contratual_sem_sigla_levanta(tmp_path):
    path = _make_xlsx(
        [['FUNÇÃO', 'CENTRAL'], ['ELET-I', 5]],
        tmp_path / 'in.xlsx',
    )
    with pytest.raises(DistribuicaoContratualMalformadaError):
        ler_xlsx_contratual(path)


def test_ler_xlsx_contratual_workbook_vazio_levanta(tmp_path):
    path = _make_xlsx([], tmp_path / 'in.xlsx')
    with pytest.raises(DistribuicaoContratualMalformadaError):
        ler_xlsx_contratual(path)


def test_ler_xlsx_contratual_strips_strings(tmp_path):
    path = _make_xlsx(
        [['  SIGLA  ', '  CENTRAL  '], ['  ELET-I  ', 5]],
        tmp_path / 'in.xlsx',
    )
    headers, data_rows, _ = ler_xlsx_contratual(path)
    assert headers == ('SIGLA', 'CENTRAL')
    assert data_rows == [('ELET-I', 5)]


def test_ler_xlsx_contratual_determinismo(tmp_path):
    path = _make_xlsx(
        [
            ['TP MO', 'FUNÇÃO', 'SIGLA', 'CENTRAL', 'BREAKDOWN PVC', 'Atual'],
            ['MOD', 'ELETRICISTA I', 'ELET-I', 5, 2, 7],
            ['MOD', 'MECANICO II',   'MEC-II', 3, 1, 4],
        ],
        tmp_path / 'in.xlsx',
    )
    primeiro = ler_xlsx_contratual(path)
    segundo = ler_xlsx_contratual(path)
    assert primeiro == segundo
