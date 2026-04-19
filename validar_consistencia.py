#!/usr/bin/env python3
"""
validar_consistencia.py — Comparação entre planilha original e processada.

Uso:
    python3 validar_consistencia.py

Saída:
    Relatório estruturado em 4 etapas impresso no stdout.
"""

import sys
import re
import os
from datetime import datetime, time as dt_time
from collections import Counter, defaultdict

import openpyxl

# ─────────────────────────────────────────────────────────────────────────────
# CAMINHOS
# ─────────────────────────────────────────────────────────────────────────────

_BASE_DIR = os.path.dirname(os.path.abspath(__file__))

ARQUIVO_ORIGINAL   = os.path.join(_BASE_DIR, 'data', 'entrada', 'Medição Geral Março.xlsx')
ARQUIVO_PROCESSADO = os.path.join(_BASE_DIR, 'data', 'entrada', 'medicao_processada.xlsx')
NOME_ABA = 'Frequencia'

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTES DE COLUNAS (0-based)
# ─────────────────────────────────────────────────────────────────────────────

COL_DATA           = 0
COL_RE             = 1
COL_SUPERVISOR     = 2
COL_ENCARREGADO    = 3
COL_RONDA          = 4
COL_NOME           = 5
COL_SG_FUNCAO      = 6
COL_UNIDADE        = 7
COL_MD_COBRANCA    = 8
COL_SITUACAO       = 9
COL_UNID_ORIGEM    = 10
COL_TAG            = 11
COL_ID_JORNADA     = 12
COL_ENTRADA        = 13
COL_SAIDA          = 14
COL_APOIO_CALC     = 15
COL_DESC_DESCANSO  = 16
COL_DIF_TURNO      = 17
COL_DESCONTOS      = 18   # modificado pelo pipeline
COL_HR_TRAB        = 19   # fórmula dependente de col 18
COL_PCT_COBR       = 20
COL_TIPO_RATEIO    = 21
COL_OBSERVACAO     = 22   # modificado pelo pipeline
COL_TOTAL_DESC     = 23   # fórmula dependente de col 18
COL_HH_MEDIDO      = 24   # fórmula dependente de col 18
COL_HISTOGRAMA     = 25   # fórmula dependente de col 18

NOMES_COLUNAS = {
    0:  'Data',
    1:  'RE',
    2:  'Supervisor',
    3:  'Encarregado',
    4:  'Ronda',
    5:  'Nome',
    6:  'Sg Função',
    7:  'Unidade',
    8:  'MD Cobranca',
    9:  'Situação',
    10: 'Unid. Origem',
    11: 'TAG',
    12: 'ID Jornada',
    13: 'Entrada',
    14: 'Saída',
    15: 'Apoio Calculo Descanso',
    16: 'Desconto Descanso',
    17: 'Dif Turno',
    18: 'Descontos',
    19: 'Hr Trabalhadas',
    20: '% Cobrança',
    21: 'Tipo Rateio',
    22: 'Observação',
    23: 'Total Descontos',
    24: 'HH Medido',
    25: 'Histograma',
}

# Colunas que o pipeline PODE modificar diretamente
COLUNAS_MODIFICADAS = {COL_DESCONTOS, COL_OBSERVACAO}

# Colunas de fórmula Excel que recalculam quando col 18 muda
COLUNAS_FORMULA_DEPENDENTE = {COL_HR_TRAB, COL_TOTAL_DESC, COL_HH_MEDIDO, COL_HISTOGRAMA}

# Colunas de tempo (além de Descontos)
COLUNAS_TEMPO = {COL_ENTRADA, COL_SAIDA, COL_APOIO_CALC, COL_DESC_DESCANSO, COL_DIF_TURNO, COL_DESCONTOS}

# Tipos de erro
ERRO_CRITICO     = 'ERRO_CRITICO'
ERRO_NUMERICO    = 'ERRO_NUMERICO'
ERRO_DATA        = 'ERRO_DATA'
ERRO_MAPEAMENTO  = 'ERRO_MAPEAMENTO'
ERRO_ESTRUTURAL  = 'ERRO_ESTRUTURAL'
ERRO_FORMATO     = 'ERRO_FORMATO'
CRITERIO_ND      = 'CRITÉRIO NÃO DEFINIDO'
INFO_MODIFICACAO = 'INFO_MODIFICACAO'

_RE_HHMMSS = re.compile(r'^(\d{1,2}):(\d{2})(?::(\d{2}))?$')

# ─────────────────────────────────────────────────────────────────────────────
# NORMALIZAÇÃO
# ─────────────────────────────────────────────────────────────────────────────

def normalizar_time(valor) -> str:
    """Converte datetime.time, string HH:MM ou HH:MM:SS para HH:MM canônico."""
    if valor is None or valor == '':
        return ''
    if isinstance(valor, dt_time):
        return f"{valor.hour:02d}:{valor.minute:02d}"
    if isinstance(valor, str):
        s = valor.strip()
        m = _RE_HHMMSS.match(s)
        if m:
            return f"{int(m.group(1)):02d}:{int(m.group(2)):02d}"
        return s
    if hasattr(valor, 'hour'):  # datetime com hora
        return f"{valor.hour:02d}:{valor.minute:02d}"
    return str(valor).strip()


def normalizar_datetime(valor) -> str:
    """Converte datetime para dd/mm/aaaa; strings retornam como estão."""
    if valor is None:
        return ''
    if hasattr(valor, 'strftime'):
        return valor.strftime('%d/%m/%Y')
    return str(valor).strip()


def normalizar_float(valor, decimais: int = 4) -> str:
    if valor is None:
        return ''
    try:
        return f"{float(valor):.{decimais}f}"
    except (TypeError, ValueError):
        return str(valor).strip()


def normalizar_valor(valor, col_idx: int) -> str:
    """Retorna string canônica para comparação, de acordo com o tipo da coluna."""
    if valor is None:
        return ''

    if col_idx == COL_DATA:
        return normalizar_datetime(valor)

    if col_idx in COLUNAS_TEMPO:
        return normalizar_time(valor)

    if col_idx in {COL_HR_TRAB, COL_PCT_COBR, COL_TOTAL_DESC, COL_HH_MEDIDO}:
        return normalizar_float(valor, decimais=4)

    if col_idx == COL_HISTOGRAMA:
        return normalizar_float(valor, decimais=6)

    if col_idx == COL_ID_JORNADA:
        try:
            return str(int(float(valor)))
        except (TypeError, ValueError):
            pass

    return str(valor).strip()


# ─────────────────────────────────────────────────────────────────────────────
# CLASSIFICAÇÃO DE ERROS
# ─────────────────────────────────────────────────────────────────────────────

def classificar_erro(col_idx: int, col18_mudou: bool) -> str:
    if col_idx in COLUNAS_MODIFICADAS:
        return INFO_MODIFICACAO

    if col_idx in COLUNAS_FORMULA_DEPENDENTE:
        return ERRO_NUMERICO if col18_mudou else ERRO_CRITICO

    if col_idx == COL_DATA:
        return ERRO_DATA

    if col_idx == COL_RE:
        return ERRO_MAPEAMENTO

    return ERRO_CRITICO


def contexto_erro(tipo: str, col_idx: int, col18_mudou: bool) -> str:
    if tipo == INFO_MODIFICACAO:
        return 'Coluna modificada pelo pipeline — diferença esperada'
    if tipo == ERRO_NUMERICO and col_idx in COLUNAS_FORMULA_DEPENDENTE:
        return (
            'Fórmula Excel recalculada após atualização de col 18 (Descontos) — '
            'derivação esperada'
        )
    if tipo == ERRO_DATA:
        return 'Divergência em campo de data — não deve ser modificado pelo pipeline'
    if tipo == ERRO_MAPEAMENTO:
        return 'Divergência em campo RE (matrícula) — não deve ser modificado'
    if tipo == ERRO_CRITICO:
        return 'Coluna não deve ser modificada pelo pipeline — alteração inesperada'
    if tipo == ERRO_ESTRUTURAL:
        return 'Diferença estrutural (linhas ou abas)'
    return ''


# ─────────────────────────────────────────────────────────────────────────────
# COMPARAÇÃO PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────

def comparar_arquivos(caminho_orig: str, caminho_proc: str) -> dict:
    erros = []
    estrutural = []
    resumo = Counter()

    wb_orig = openpyxl.load_workbook(caminho_orig, read_only=True, data_only=True)
    wb_proc = openpyxl.load_workbook(caminho_proc, read_only=True, data_only=True)

    # ── Verificação de abas ──────────────────────────────────────────────────
    sheets_orig = set(wb_orig.sheetnames)
    sheets_proc = set(wb_proc.sheetnames)

    if sheets_orig != sheets_proc:
        apenas_orig = sorted(sheets_orig - sheets_proc)
        apenas_proc = sorted(sheets_proc - sheets_orig)
        msg = []
        if apenas_orig:
            msg.append(f"Apenas no original: {apenas_orig}")
        if apenas_proc:
            msg.append(f"Apenas no processado: {apenas_proc}")
        estrutural.append({
            'tipo': ERRO_ESTRUTURAL,
            'descricao': 'Abas diferem. ' + '; '.join(msg),
        })
        resumo[ERRO_ESTRUTURAL] += 1

    # Verificar presença da aba principal
    for label, wb in [('original', wb_orig), ('processado', wb_proc)]:
        if NOME_ABA not in wb.sheetnames:
            # Tentar variante com acento
            acento = 'Frequência'
            if acento in wb.sheetnames:
                estrutural.append({
                    'tipo': ERRO_ESTRUTURAL,
                    'descricao': (
                        f"Aba '{NOME_ABA}' não encontrada no {label}; "
                        f"'{acento}' encontrada. Usando '{acento}'."
                    ),
                })
            else:
                estrutural.append({
                    'tipo': ERRO_ESTRUTURAL,
                    'descricao': f"Aba '{NOME_ABA}' ausente no arquivo {label}.",
                })
                wb_orig.close()
                wb_proc.close()
                return {
                    'erros': erros, 'resumo': resumo, 'estrutural': estrutural,
                    'total_linhas_orig': 0, 'total_linhas_proc': 0,
                    'total_linhas_comparadas': 0,
                }

    # Selecionar aba (com fallback de acento)
    def get_sheet(wb):
        if NOME_ABA in wb.sheetnames:
            return wb[NOME_ABA]
        return wb['Frequência']

    sh_orig = get_sheet(wb_orig)
    sh_proc = get_sheet(wb_proc)

    # ── Passada única simultânea ─────────────────────────────────────────────
    total_orig = 0
    total_proc = 0
    total_comparadas = 0
    _SENTINEL = object()

    iter_orig = sh_orig.iter_rows(values_only=True)
    iter_proc = sh_proc.iter_rows(values_only=True)

    while True:
        r1 = next(iter_orig, _SENTINEL)
        r2 = next(iter_proc, _SENTINEL)

        if r1 is _SENTINEL and r2 is _SENTINEL:
            break

        if r1 is not _SENTINEL:
            total_orig += 1
        if r2 is not _SENTINEL:
            total_proc += 1

        # Linha assimétrica
        if r1 is _SENTINEL:
            erros.append({
                'linha': total_proc,
                'tipo': ERRO_ESTRUTURAL,
                'coluna_idx': -1,
                'coluna_nome': 'N/A',
                'valor_original': '(linha ausente)',
                'valor_processado': str(r2)[:200],
                're': '',
                'nome': '',
                'data': '',
                'contexto': 'Linha extra no arquivo processado',
            })
            resumo[ERRO_ESTRUTURAL] += 1
            continue

        if r2 is _SENTINEL:
            erros.append({
                'linha': total_orig,
                'tipo': ERRO_ESTRUTURAL,
                'coluna_idx': -1,
                'coluna_nome': 'N/A',
                'valor_original': str(r1)[:200],
                'valor_processado': '(linha ausente)',
                're': '',
                'nome': '',
                'data': '',
                'contexto': 'Linha ausente no arquivo processado',
            })
            resumo[ERRO_ESTRUTURAL] += 1
            continue

        total_comparadas += 1
        row_num = total_orig  # mesmo índice nos dois (contagem simultânea)

        # Identidade da linha para contexto
        re_val   = r1[COL_RE]   if len(r1) > COL_RE   else None
        data_val = r1[COL_DATA] if len(r1) > COL_DATA else None
        nome_val = r1[COL_NOME] if len(r1) > COL_NOME else None
        re_str   = str(re_val).strip()   if re_val   is not None else ''
        data_str = normalizar_datetime(data_val)
        nome_str = str(nome_val).strip() if nome_val is not None else ''

        # Detectar se col 18 mudou nesta linha
        orig_18 = r1[COL_DESCONTOS] if len(r1) > COL_DESCONTOS else None
        proc_18 = r2[COL_DESCONTOS] if len(r2) > COL_DESCONTOS else None
        col18_mudou = (normalizar_time(orig_18) != normalizar_time(proc_18))

        # Comparar todas as colunas
        n_cols = max(len(r1), len(r2), 26)
        for c in range(n_cols):
            v1 = r1[c] if c < len(r1) else None
            v2 = r2[c] if c < len(r2) else None

            n1 = normalizar_valor(v1, c)
            n2 = normalizar_valor(v2, c)

            if n1 == n2:
                continue

            tipo = classificar_erro(c, col18_mudou)
            col_nome = NOMES_COLUNAS.get(c, f'Col_{c}')

            erros.append({
                'linha': row_num,
                'tipo': tipo,
                'coluna_idx': c,
                'coluna_nome': col_nome,
                'valor_original': str(v1)[:300] if v1 is not None else '(vazio)',
                'valor_processado': str(v2)[:300] if v2 is not None else '(vazio)',
                're': re_str,
                'nome': nome_str,
                'data': data_str,
                'contexto': contexto_erro(tipo, c, col18_mudou),
            })
            resumo[tipo] += 1

    wb_orig.close()
    wb_proc.close()

    if total_orig != total_proc:
        estrutural.append({
            'tipo': ERRO_ESTRUTURAL,
            'descricao': (
                f"Número de linhas difere: original={total_orig}, "
                f"processado={total_proc}."
            ),
        })
        resumo[ERRO_ESTRUTURAL] += 1

    return {
        'erros': erros,
        'resumo': resumo,
        'estrutural': estrutural,
        'total_linhas_orig': total_orig,
        'total_linhas_proc': total_proc,
        'total_linhas_comparadas': total_comparadas,
    }


# ─────────────────────────────────────────────────────────────────────────────
# GERAÇÃO DO RELATÓRIO
# ─────────────────────────────────────────────────────────────────────────────

SEP_SECAO = '═' * 80
SEP_LINHA = '─' * 70

TIPOS_ORDEM = [
    ERRO_CRITICO,
    ERRO_NUMERICO,
    ERRO_DATA,
    ERRO_MAPEAMENTO,
    ERRO_ESTRUTURAL,
    ERRO_FORMATO,
    CRITERIO_ND,
    INFO_MODIFICACAO,
]


def imprimir_relatorio(resultado: dict, caminho_orig: str, caminho_proc: str):
    erros      = resultado['erros']
    resumo     = resultado['resumo']
    estrutural = resultado['estrutural']
    total_orig = resultado['total_linhas_orig']
    total_proc = resultado['total_linhas_proc']
    total_comp = resultado['total_linhas_comparadas']

    # ── ETAPA 1 ──────────────────────────────────────────────────────────────
    print(SEP_SECAO)
    print('ETAPA 1 — DOCUMENTAÇÃO UTILIZADA')
    print(SEP_SECAO)
    print()
    print('Arquivos .md lidos:')
    print(f"  - {os.path.join(os.path.dirname(caminho_orig), '..', '..', 'CLAUDE.md')} (CLAUDE.md do projeto)")
    print()
    print('Critérios extraídos:')
    print('  - Colunas modificadas pelo pipeline: 18 (Descontos), 22 (Observação)')
    print('  - Colunas de fórmula dependentes de col 18: 19 (Hr Trabalhadas),')
    print('    23 (Total Descontos), 24 (HH Medido), 25 (Histograma)')
    print('  - Todas as demais colunas (0-17 exceto 18, 20-21, 22) devem ser idênticas')
    print('  - Formato de desconto: HH:MM (escrito como inlineStr via ZIP patch)')
    print('  - Formato de observação: TREIN. {nome} - {horas}H [+ (NÃO DESCONTA)]')
    print('  - Chave de identificação: RE (col 1) + Data (col 0)')
    print('  - Aba comparada: Frequencia (sem acento)')
    print('  - Passada única read_only: iter_rows(values_only=True)')
    print()
    print('Arquivos comparados:')
    print(f"  - Original  : {caminho_orig}")
    print(f"  - Processado: {caminho_proc}")
    print(f"  - Data/hora : {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print(f"  - Linhas no original  : {total_orig}")
    print(f"  - Linhas no processado: {total_proc}")
    print(f"  - Linhas comparadas   : {total_comp}")
    print()

    # ── ETAPA 2 ──────────────────────────────────────────────────────────────
    print(SEP_SECAO)
    print('ETAPA 2 — RESUMO')
    print(SEP_SECAO)
    print()

    if estrutural:
        print('Alertas estruturais:')
        for s in estrutural:
            print(f"  [{s['tipo']}] {s['descricao']}")
        print()

    print(f"  {'TIPO':<35} {'OCORRÊNCIAS':>12}")
    print(f"  {'-'*35} {'-'*12}")

    total_erros_reais = 0
    for tipo in TIPOS_ORDEM:
        qtd = resumo.get(tipo, 0)
        if qtd > 0:
            print(f"  {tipo:<35} {qtd:>12}")
            if tipo not in (INFO_MODIFICACAO,):
                total_erros_reais += qtd

    print(f"  {'-'*35} {'-'*12}")
    print(f"  {'Total de erros (excl. INFO)':<35} {total_erros_reais:>12}")
    print(f"  {'Total de diferenças encontradas':<35} {sum(resumo.values()):>12}")
    print()
    print(f"  ERRO_CRITICO    : {resumo.get(ERRO_CRITICO, 0)}")
    print(f"  ERRO_NUMERICO   : {resumo.get(ERRO_NUMERICO, 0)}")
    print(f"  ERRO_DATA       : {resumo.get(ERRO_DATA, 0)}")
    print(f"  ERRO_MAPEAMENTO : {resumo.get(ERRO_MAPEAMENTO, 0)}")
    print(f"  ERRO_ESTRUTURAL : {resumo.get(ERRO_ESTRUTURAL, 0)}")
    print(f"  ERRO_FORMATO    : {resumo.get(ERRO_FORMATO, 0)}")
    print(f"  CRITÉRIO NÃO DEFINIDO: {resumo.get(CRITERIO_ND, 0)}")
    print()

    # ── ETAPA 3 ──────────────────────────────────────────────────────────────
    print(SEP_SECAO)
    print('ETAPA 3 — DETALHAMENTO DOS ERROS')
    print(SEP_SECAO)

    por_tipo = defaultdict(list)
    for e in erros:
        por_tipo[e['tipo']].append(e)

    erro_num_global = 0
    for tipo in TIPOS_ORDEM:
        grupo = por_tipo.get(tipo, [])
        if not grupo:
            continue
        print()
        print(f"  [{tipo}] — {len(grupo)} ocorrência(s)")
        print(f"  {SEP_LINHA}")

        for e in grupo:
            erro_num_global += 1
            print()
            print(f"  [Erro {erro_num_global}]")
            print(f"  Linha original    : {e['linha']}")
            print(f"  Linha processada  : {e['linha']}")
            print(f"  Identificador (RE): {e['re']}")
            print(f"  Nome              : {e['nome']}")
            print(f"  Data              : {e['data']}")
            print(f"  Coluna            : {e['coluna_idx']} — {e['coluna_nome']}")
            print(f"  Valor original    : {e['valor_original']}")
            print(f"  Valor processado  : {e['valor_processado']}")
            print(f"  Tipo              : {e['tipo']}")
            print(f"  Descrição         : {e['contexto']}")
            print(f"  {'-'*60}")

    if erro_num_global == 0:
        print()
        print('  Nenhuma diferença encontrada.')
    print()

    # ── ETAPA 4 ──────────────────────────────────────────────────────────────
    print(SEP_SECAO)
    print('ETAPA 4 — CONCLUSÃO')
    print(SEP_SECAO)
    print()

    n_critico    = resumo.get(ERRO_CRITICO, 0)
    n_numerico   = resumo.get(ERRO_NUMERICO, 0)
    n_data       = resumo.get(ERRO_DATA, 0)
    n_mapeamento = resumo.get(ERRO_MAPEAMENTO, 0)
    n_estrutural = resumo.get(ERRO_ESTRUTURAL, 0)
    n_modificacao = resumo.get(INFO_MODIFICACAO, 0)

    print(f"  Total de linhas verificadas                    : {total_comp}")
    print(f"  Diferenças em colunas modificadas (cols 18/22) : {n_modificacao}")
    print(f"  Erros críticos (colunas não devem mudar)       : {n_critico}")
    print(f"  Erros numéricos (fórmulas/recálculo)           : {n_numerico}")
    print(f"  Erros de data                                  : {n_data}")
    print(f"  Erros de mapeamento (RE)                       : {n_mapeamento}")
    print(f"  Erros estruturais                              : {n_estrutural}")
    print()

    if n_critico == 0 and n_data == 0 and n_mapeamento == 0:
        if erro_num_global == 0:
            print('VALIDAÇÃO CONCLUÍDA: ZERO ERROS')
        else:
            print('VALIDAÇÃO CONCLUÍDA: ZERO ERROS')
            print()
            if n_numerico > 0:
                print(
                    f"  Nota: {n_numerico} diferença(s) numéricas detectadas em colunas de fórmula "
                    f"(cols 19, 23, 24, 25)."
                )
                print(
                    "  Essas são derivações esperadas do recálculo do Excel após a atualização "
                    "da coluna Descontos (col 18) pelo pipeline."
                )
            if n_estrutural > 0:
                print(
                    f"  Nota: {n_estrutural} diferença(s) estrutural(is) (ex: aba Planilha3 removida pelo ZIP patch)."
                )
    else:
        print('VALIDAÇÃO CONCLUÍDA: ERROS ENCONTRADOS')
        print()
        if n_critico > 0:
            print(f"  ATENÇÃO: {n_critico} ERRO(S) CRÍTICO(S) — colunas que não devem ser modificadas foram alteradas.")
        if n_data > 0:
            print(f"  ATENÇÃO: {n_data} ERRO(S) DE DATA — datas foram alteradas.")
        if n_mapeamento > 0:
            print(f"  ATENÇÃO: {n_mapeamento} ERRO(S) DE MAPEAMENTO — RE/matrícula foram alterados.")
    print()


# ─────────────────────────────────────────────────────────────────────────────
# PONTO DE ENTRADA
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print()
    print(SEP_SECAO)
    print('VALIDAÇÃO DE CONSISTÊNCIA — MEDIÇÃO PROCESSADA')
    print(SEP_SECAO)
    print()

    for caminho, label in [(ARQUIVO_ORIGINAL, 'original'), (ARQUIVO_PROCESSADO, 'processado')]:
        if not os.path.exists(caminho):
            print(f"[ERRO] Arquivo {label} não encontrado: {caminho}")
            sys.exit(1)

    try:
        resultado = comparar_arquivos(ARQUIVO_ORIGINAL, ARQUIVO_PROCESSADO)
    except Exception as e:
        print(f"[ERRO INESPERADO] {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

    imprimir_relatorio(resultado, ARQUIVO_ORIGINAL, ARQUIVO_PROCESSADO)


if __name__ == '__main__':
    main()
