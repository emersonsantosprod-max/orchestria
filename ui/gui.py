import os
import sys
import platform
import ctypes
import threading
import tkinter as tk
import tkinter.font as tkfont
from pathlib import Path
from tkinter import filedialog

import customtkinter as ctk

# ---------------------------------------------------------------------------
# Path setup
# ---------------------------------------------------------------------------

def get_base_path():
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS
    return os.path.abspath(".")


base_path = get_base_path()
app_path = os.path.join(base_path, "app")
if app_path not in sys.path:
    sys.path.append(app_path)

try:
    import openpyxl  # noqa: F401
except ImportError:
    pass

from app.pipeline import processar, salvar_relatorio_inconsistencias
from app import db
from app.validar_distribuicao import validar, gerar_relatorio, _salvar_relatorio
from app.validar_horas import (
    validar as _validar_hr,
    gerar_relatorio as _gerar_relatorio_hr,
    _salvar_relatorio as _salvar_relatorio_hr,
)

# ---------------------------------------------------------------------------
# Design tokens (Manserv brand)
# ---------------------------------------------------------------------------

_CHUMBO      = "#232323"
_CHUMBO_2    = "#111111"
_LARANJA     = "#ff460a"
_LARANJA_HV  = "#e2360e"
_CONTEUDO    = "#ededed"
_PAINEL      = "#ffffff"
_TXT_INV     = "#ffffff"
_TXT_NAV     = "#e5e5e5"

# Column indices — Frequencia sheet (from app/cli/validar_consist.py)
_COL_DATA    = 0
_COL_RE      = 1
_COL_HR_TRAB = 19

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def selecionar_arquivo(titulo):
    return filedialog.askopenfilename(
        title=titulo,
        filetypes=[("Arquivos Excel", "*.xlsx")],
    )


def _todos_botoes():
    return [botao_lancar, botao_ferias, botao_atestado, botao_validar, botao_validar_hr]


def _desabilitar_botoes():
    for b in _todos_botoes():
        b.configure(state="disabled")


def _habilitar_botoes():
    janela.after(0, lambda: [b.configure(state="normal") for b in _todos_botoes()])


# ---------------------------------------------------------------------------
# Handlers
# ---------------------------------------------------------------------------

def _executar_fluxo(titulo_log, prompts, montar_kwargs):
    _desabilitar_botoes()
    area_saida.delete("1.0", "end")

    caminhos = {}
    for label, chave in prompts:
        caminho = selecionar_arquivo(label)
        if not caminho:
            imprimir_log(f"Operação cancelada: {label} não foi selecionada.\n")
            _habilitar_botoes()
            return
        caminhos[chave] = caminho

    imprimir_log(f"Iniciando {titulo_log}...\n")

    def tarefa():
        conn = db.conectar()
        try:
            db.popular_bd_se_vazio(conn)
            imprimir_log("Fase 1/3: Lendo arquivos (modo otimizado)...\n")
            resultado = processar(
                **montar_kwargs(caminhos),
                conn=conn,
                validar_distribuicao=False,
            )
            imprimir_log("Fase 2/3: Processando regras de negócio...\n")
            imprimir_log("Fase 3/3: Gravando resultados no Excel (isso pode demorar)...\n")
            mostrar_resultado(resultado)
        except Exception as e:
            imprimir_log(f"\n[ERRO] {str(e)}")
        finally:
            conn.close()
            _habilitar_botoes()

    threading.Thread(target=tarefa, daemon=True).start()


def iniciar_lancamento():
    _executar_fluxo(
        titulo_log="processamento de treinamentos",
        prompts=[
            ("1. Selecione a planilha de Medição Destino", 'medicao'),
            ("2. Selecione a planilha de Treinamentos Realizados", 'treinamentos'),
            ("3. Selecione a Base de Treinamentos (Classificação)", 'classificacao'),
        ],
        montar_kwargs=lambda c: dict(
            caminho_medicao=c['medicao'],
            caminho_treinamentos=c['treinamentos'],
            caminho_classificacao=c['classificacao'],
        ),
    )


def iniciar_ferias():
    _executar_fluxo(
        titulo_log="processamento de férias",
        prompts=[
            ("1. Selecione a planilha de Medição Destino", 'medicao'),
            ("2. Selecione o Relatório Geral de Férias", 'ferias'),
            ("3. Selecione a Base de Cobrança (SgFunção → Categoria)", 'base_cobranca'),
        ],
        montar_kwargs=lambda c: dict(
            caminho_medicao=c['medicao'],
            caminho_ferias=c['ferias'],
            caminho_base_cobranca=c['base_cobranca'],
        ),
    )


def iniciar_atestado():
    _executar_fluxo(
        titulo_log="processamento de atestados médicos",
        prompts=[
            ("1. Selecione a planilha de Medição Destino", 'medicao'),
            ("2. Selecione o arquivo de Atestados Médicos", 'atestado'),
        ],
        montar_kwargs=lambda c: dict(
            caminho_medicao=c['medicao'],
            caminho_atestado=c['atestado'],
        ),
    )


def iniciar_validacao():
    _desabilitar_botoes()
    area_saida.delete("1.0", "end")

    conn = db.conectar()
    db.popular_bd_se_vazio(conn)
    registros = db.obter_registro_arquivos(conn)
    conn.close()

    caminho_bd = None
    caminho_medicao = None

    if 'bd' not in registros:
        imprimir_log("BD não registrado — solicitando arquivo...\n")
        caminho_bd = selecionar_arquivo("Selecione o arquivo de BD de Distribuição Contratual")
        if not caminho_bd:
            imprimir_log("Operação cancelada: BD não foi selecionado.\n")
            _habilitar_botoes()
            return

    if 'medicao' not in registros:
        imprimir_log("Medição não registrada — solicitando arquivo...\n")
        caminho_medicao = selecionar_arquivo("Selecione o arquivo de Medição Frequência")
        if not caminho_medicao:
            imprimir_log("Operação cancelada: Medição não foi selecionada.\n")
            _habilitar_botoes()
            return

    imprimir_log("Executando validação...\n")

    def tarefa():
        try:
            conn = db.conectar()
            avisos_import = []

            if caminho_bd:
                imprimir_log("Registrando BD...\n")
                db.registrar_bd(caminho_bd, conn)

            if caminho_medicao:
                imprimir_log("Registrando Medição...\n")
                avisos = db.registrar_medicao(caminho_medicao, conn)
                avisos_import.extend(avisos)

            registros_atuais = db.obter_registro_arquivos(conn)
            bd_records       = db.obter_bd(conn)
            medicao_records  = db.obter_medicao(conn)
            conn.close()

            inconsistencias = validar(bd_records, medicao_records)

            bd_pares = {(r['funcao'], r['md_cobranca']) for r in bd_records}
            datas    = {r['data'] for r in medicao_records}

            conteudo = gerar_relatorio(
                inconsistencias, registros_atuais,
                n_pares_bd=len(bd_pares),
                n_datas=len(datas),
                avisos_import=avisos_import,
            )
            caminho_rel = _salvar_relatorio(conteudo)

            imprimir_log(f"Relatório gerado em: {caminho_rel}\n")
            imprimir_log(f"Total de inconsistências: {len(inconsistencias)}\n")
            for av in avisos_import:
                imprimir_log(f"[AVISO] {av}\n")

        except Exception as e:
            imprimir_log(f"\n[ERRO] {str(e)}\n")
        finally:
            _habilitar_botoes()

    threading.Thread(target=tarefa, daemon=True).start()


def _ler_medicao_hr(path: Path):
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        ws = wb['Frequencia'] if 'Frequencia' in wb.sheetnames else wb['Frequência']
        registros = []
        n = 0
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            if row is None or all(c is None for c in row):
                continue
            n += 1
            mat_val = row[_COL_RE]      if len(row) > _COL_RE      else None
            dat_val = row[_COL_DATA]    if len(row) > _COL_DATA    else None
            hr_val  = row[_COL_HR_TRAB] if len(row) > _COL_HR_TRAB else None
            matricula = str(mat_val).strip() if mat_val is not None else ''
            if hasattr(dat_val, 'strftime'):
                data_str = dat_val.strftime('%d/%m/%Y')
            else:
                data_str = str(dat_val).strip() if dat_val is not None else ''
            hr_trab = float(hr_val) if hr_val is not None else None
            registros.append({'matricula': matricula, 'data': data_str, 'hr_trabalhadas': hr_trab})
    finally:
        wb.close()
    return registros, n


def iniciar_validar_hr():
    _desabilitar_botoes()
    area_saida.delete("1.0", "end")

    caminho = selecionar_arquivo("Selecione a planilha de Medição")
    if not caminho:
        imprimir_log("Operação cancelada: arquivo não foi selecionado.\n")
        _habilitar_botoes()
        return

    imprimir_log("Validando horas trabalhadas...\n")

    def tarefa():
        try:
            registros, n_linhas = _ler_medicao_hr(Path(caminho))
            inconsistencias = _validar_hr(registros)
            conteudo = _gerar_relatorio_hr(inconsistencias, str(Path(caminho).resolve()), n_linhas)
            caminho_rel = _salvar_relatorio_hr(conteudo)
            imprimir_log(f"Relatório gerado em: {caminho_rel}\n")
            imprimir_log(f"Total de inconsistências: {len(inconsistencias)}\n")
        except Exception as e:
            imprimir_log(f"\n[ERRO] {str(e)}\n")
        finally:
            _habilitar_botoes()

    threading.Thread(target=tarefa, daemon=True).start()


def mostrar_resultado(resultado):
    processados    = resultado.processados
    atualizados    = resultado.atualizados
    ferias_proc    = resultado.ferias_processadas
    ferias_atu     = resultado.ferias_atualizadas
    atestados_proc = resultado.atestados_processados
    atestados_atu  = resultado.atestados_atualizados
    inconsistencias = resultado.inconsistencias
    caminho_saida   = resultado.caminho_saida

    log = (
        "Processamento concluído com sucesso.\n"
        "--------------------------------------------------\n"
        f"Treinamentos processados: {processados}\n"
        f"Treinamentos atualizados: {atualizados}\n"
        f"Férias processadas:       {ferias_proc}\n"
        f"Linhas com férias:        {ferias_atu}\n"
        f"Atestados processados:    {atestados_proc}\n"
        f"Linhas com atestado:      {atestados_atu}\n"
    )
    nome_arquivo = os.path.basename(caminho_saida)
    log += f"Arquivo gerado:         {nome_arquivo}\n"
    log += f"Inconsistências totais: {len(inconsistencias)}\n"

    if inconsistencias:
        log += "\n--- LISTA DE INCONSISTÊNCIAS (preview: primeiros 10) ---\n"
        for inc in inconsistencias[:10]:
            mat  = inc.matricula or '-'
            data = inc.data or '-'
            erro = inc.erro or 'erro desconhecido'
            log += f"{mat} | {data} | {erro}\n"
        if len(inconsistencias) > 10:
            log += f"... ({len(inconsistencias) - 10} inconsistências adicionais no relatório)\n"

        selected_dir = filedialog.askdirectory(
            title="Selecione o diretório para salvar o relatório de inconsistências"
        )
        output_dir = selected_dir if selected_dir else os.path.dirname(caminho_saida)
        log += "\nExportando relatório de inconsistências...\n"
        caminho_relatorio = salvar_relatorio_inconsistencias(output_dir, inconsistencias)
        if caminho_relatorio:
            log += f"Relatório salvo: {caminho_relatorio}\n"

    imprimir_log(log)


def imprimir_log(texto):
    def inserir():
        area_saida.insert("end", texto)
        area_saida.see("end")
    janela.after(0, inserir)


# ---------------------------------------------------------------------------
# Sidebar navigation
# ---------------------------------------------------------------------------

def _ativar_aba(tab_id):
    if tab_id == 'lancar':
        nav_lancar.configure(fg_color=_LARANJA, text_color=_TXT_INV, font=_fonte_nav_ativo)
        nav_validacao.configure(fg_color="transparent", text_color=_TXT_NAV, font=_fonte_nav)
        frame_validacao.pack_forget()
        frame_lancar.pack(padx=16, pady=12, fill="x")
    else:
        nav_lancar.configure(fg_color="transparent", text_color=_TXT_NAV, font=_fonte_nav)
        nav_validacao.configure(fg_color=_LARANJA, text_color=_TXT_INV, font=_fonte_nav_ativo)
        frame_lancar.pack_forget()
        frame_validacao.pack(padx=16, pady=12, fill="x")


# ---------------------------------------------------------------------------
# Window
# ---------------------------------------------------------------------------

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

janela = ctk.CTk()
janela.title("Automação de Medição")
janela.geometry("860x560")
janela.minsize(700, 480)

# Load IBM Plex Sans from design/fonts/ on Windows
if platform.system() == "Windows":
    for _fn in ("IBMPlexSans-Regular.ttf", "IBMPlexSans-SemiBold.ttf"):
        _fp = os.path.join(base_path, "design", "fonts", _fn)
        if os.path.exists(_fp):
            try:
                ctypes.windll.gdi32.AddFontResourceExW(_fp, 0x10, 0)
            except Exception:
                pass

janela.update_idletasks()
_familia = "IBM Plex Sans" if "IBM Plex Sans" in tkfont.families() else "Segoe UI"
_w = janela.winfo_width()
_h = janela.winfo_height()
_x = (janela.winfo_screenwidth() - _w) // 2
_y = (janela.winfo_screenheight() - _h) // 2
janela.geometry(f"{_w}x{_h}+{_x}+{_y}")

_fonte_nav      = ctk.CTkFont(family=_familia, size=13)
_fonte_nav_ativo = ctk.CTkFont(family=_familia, size=13, weight="bold")
_fonte_heading  = ctk.CTkFont(family=_familia, size=14, weight="bold")
_fonte_botao    = ctk.CTkFont(family=_familia, size=12)
_fonte_label    = ctk.CTkFont(family=_familia, size=11)
_fonte_log      = ctk.CTkFont(family="Courier New", size=10)

janela.grid_columnconfigure(1, weight=1)
janela.grid_rowconfigure(0, weight=1)

# ---------------------------------------------------------------------------
# Sidebar
# ---------------------------------------------------------------------------

sidebar = ctk.CTkFrame(janela, width=200, corner_radius=0, fg_color=_CHUMBO)
sidebar.grid(row=0, column=0, sticky="nsew")
sidebar.grid_propagate(False)
sidebar.grid_columnconfigure(0, weight=1)
sidebar.grid_rowconfigure(5, weight=1)

ctk.CTkLabel(
    sidebar,
    text="Automação de\nMedição",
    font=_fonte_heading,
    text_color=_TXT_INV,
    justify="left",
    anchor="w",
).grid(row=0, column=0, padx=18, pady=(20, 6), sticky="w")

tk.Frame(sidebar, height=1, bg="#3a3a3a").grid(
    row=1, column=0, padx=18, pady=(2, 10), sticky="ew"
)

ctk.CTkLabel(
    sidebar,
    text="MÓDULOS",
    font=ctk.CTkFont(family=_familia, size=10, weight="bold"),
    text_color="#6a6a6a",
    anchor="w",
).grid(row=2, column=0, padx=18, pady=(0, 4), sticky="w")

nav_lancar = ctk.CTkButton(
    sidebar,
    text="Lançar",
    font=_fonte_nav_ativo,
    fg_color=_LARANJA,
    hover_color=_LARANJA_HV,
    text_color=_TXT_INV,
    anchor="w",
    corner_radius=4,
    height=36,
    command=lambda: _ativar_aba('lancar'),
)
nav_lancar.grid(row=3, column=0, padx=8, pady=2, sticky="ew")

nav_validacao = ctk.CTkButton(
    sidebar,
    text="Validação",
    font=_fonte_nav,
    fg_color="transparent",
    hover_color=_CHUMBO_2,
    text_color=_TXT_NAV,
    anchor="w",
    corner_radius=4,
    height=36,
    command=lambda: _ativar_aba('validacao'),
)
nav_validacao.grid(row=4, column=0, padx=8, pady=2, sticky="ew")

ctk.CTkLabel(
    sidebar,
    text="Manserv · 2026",
    font=ctk.CTkFont(family=_familia, size=10),
    text_color="#555555",
    anchor="w",
).grid(row=6, column=0, padx=18, pady=(0, 12), sticky="sw")

# ---------------------------------------------------------------------------
# Content area
# ---------------------------------------------------------------------------

content = ctk.CTkFrame(janela, corner_radius=0, fg_color=_CONTEUDO)
content.grid(row=0, column=1, sticky="nsew")

painel_botoes = ctk.CTkFrame(content, corner_radius=0, fg_color=_PAINEL)
painel_botoes.pack(fill="x")

# Lançar button group
frame_lancar = ctk.CTkFrame(painel_botoes, fg_color=_PAINEL, corner_radius=0)

botao_lancar = ctk.CTkButton(
    frame_lancar,
    text="Lançar treinamentos",
    font=_fonte_botao,
    fg_color=_LARANJA,
    hover_color=_LARANJA_HV,
    text_color=_TXT_INV,
    corner_radius=4,
    height=36,
    command=iniciar_lancamento,
)
botao_lancar.pack(side="left", padx=(0, 8))

botao_ferias = ctk.CTkButton(
    frame_lancar,
    text="Lançar férias",
    font=_fonte_botao,
    fg_color=_LARANJA,
    hover_color=_LARANJA_HV,
    text_color=_TXT_INV,
    corner_radius=4,
    height=36,
    command=iniciar_ferias,
)
botao_ferias.pack(side="left", padx=(0, 8))

botao_atestado = ctk.CTkButton(
    frame_lancar,
    text="Lançar atestados",
    font=_fonte_botao,
    fg_color=_LARANJA,
    hover_color=_LARANJA_HV,
    text_color=_TXT_INV,
    corner_radius=4,
    height=36,
    command=iniciar_atestado,
)
botao_atestado.pack(side="left")

# Validação button group
frame_validacao = ctk.CTkFrame(painel_botoes, fg_color=_PAINEL, corner_radius=0)

botao_validar = ctk.CTkButton(
    frame_validacao,
    text="Validar distribuição",
    font=_fonte_botao,
    fg_color=_LARANJA,
    hover_color=_LARANJA_HV,
    text_color=_TXT_INV,
    corner_radius=4,
    height=36,
    command=iniciar_validacao,
)
botao_validar.pack(side="left", padx=(0, 8))

botao_validar_hr = ctk.CTkButton(
    frame_validacao,
    text="Validar Hr Trabalhadas",
    font=_fonte_botao,
    fg_color=_LARANJA,
    hover_color=_LARANJA_HV,
    text_color=_TXT_INV,
    corner_radius=4,
    height=36,
    command=iniciar_validar_hr,
)
botao_validar_hr.pack(side="left")

# Show Lançar tab by default
frame_lancar.pack(padx=16, pady=12, fill="x")

ctk.CTkLabel(
    content,
    text="Log de execução",
    font=_fonte_label,
    text_color="#5a5a5a",
    anchor="w",
).pack(padx=16, pady=(8, 4), anchor="w")

area_saida = ctk.CTkTextbox(
    content,
    font=_fonte_log,
    fg_color="#f4f4f4",
    text_color=_CHUMBO,
    corner_radius=4,
    border_width=1,
    border_color="#d9d9d9",
    wrap="word",
)
area_saida.pack(fill="both", expand=True, padx=16, pady=(0, 16))


if __name__ == "__main__":
    try:
        janela.mainloop()
    except Exception as e:
        caminho_fatal = os.path.join(str(Path.home()), "Downloads", "erro_fatal_automacao.txt")
        with open(caminho_fatal, "w", encoding="utf-8") as f:
            import traceback
            f.write(f"ERRO FATAL QUE FECHOU A TELA:\n\n{str(e)}\n\n")
            f.write(traceback.format_exc())
        sys.exit(1)
